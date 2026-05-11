#!/usr/bin/env python3
"""
Gera planilha XLSX para controle do álbum Panini FIFA World Cup 2026 (980 figurinhas base).
Fonte dos códigos: checklist público (ordem do álbum por grupos).

Para dashboard + estatísticas avançadas, use: generate_xlsx_v2.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from wc2026_album_core import CHECKLIST_DEFAULT, COCA_COLA, EXTRAS, foil_type as foil_type_core, load_album

DIR = Path(__file__).resolve().parent
OUT = DIR / "FIFA-World-Cup-2026-Panini-Controle.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1E3A5F")
THIN = Side(style="thin", color="CCCCCC")


def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()


def add_tenho_validation(ws, col_letter: str, start_row: int, end_row: int):
    dv = DataValidation(
        type="list",
        formula1='"Sim,Não,Falta trocar"',
        allow_blank=True,
    )
    dv.error = "Escolha Sim, Não ou Falta trocar."
    dv.errorTitle = "Valor inválido"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def autosize_columns(ws, max_width: int = 52):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        maxlen = 0
        for cell in col:
            if cell.value is None:
                continue
            maxlen = max(maxlen, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_width, maxlen + 2)


def main() -> None:
    intro, museum, teams_meta = load_album(CHECKLIST_DEFAULT)

    wb = Workbook()

    # --- Resumo ---
    ws0 = wb.active
    ws0.title = "Resumo"
    ws0.merge_cells("A1:D1")
    ws0["A1"] = "Copa do Mundo FIFA 2026 — Panini | Controle de figurinhas"
    ws0["A1"].font = TITLE_FONT
    ws0["A1"].alignment = Alignment(horizontal="center")

    notes = (
        "Este arquivo inclui as 980 figurinhas do álbum base (ordem oficial Panini), "
        "as 12 Coca-Cola (promo EUA), e as 20 figurinhas Extra (internacional, sem número no álbum). "
        "Marque a coluna «Tenho» em cada secção. Duplicatas: quantas repetidas tem para troca."
    )
    ws0.merge_cells("A3:D6")
    ws0["A3"] = notes
    ws0["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    ws0["A8"] = "Métrica"
    ws0["B8"] = "Valor"
    ws0["A8"].font = Font(bold=True)
    ws0["B8"].font = Font(bold=True)

    metrics = [
        ("Figurinhas base no álbum", 980),
        ("— Página inicial (logo + FWC1–8)", len(intro)),
        ("— Selecções (48 × 20)", 960),
        ("— Museu FIFA (FWC9–19)", len(museum)),
        ("Promo Coca-Cola (fora dos pacotes)", len(COCA_COLA)),
        ("Extras internacionais (sem número)", len(EXTRAS)),
        ("Total de itens para acompanhar", 980 + len(COCA_COLA) + len(EXTRAS)),
    ]
    for r, (label, val) in enumerate(metrics, start=9):
        ws0.cell(row=r, column=1, value=label)
        ws0.cell(row=r, column=2, value=val)

    r0 = 17
    ws0.cell(row=r0, column=1, value="Figurinhas base marcadas como «Sim»")
    ws0.cell(row=r0, column=2, value="=COUNTIF('Página Inicial'!E:E,\"Sim\")+COUNTIF('Seleções'!H:H,\"Sim\")+COUNTIF('Museu FIFA'!E:E,\"Sim\")")

    ws0.cell(row=r0 + 1, column=1, value="Coca-Cola marcadas")
    ws0.cell(row=r0 + 1, column=2, value="=COUNTIF('Coca-Cola'!E:E,\"Sim\")")

    ws0.cell(row=r0 + 2, column=1, value="Extras marcadas")
    ws0.cell(row=r0 + 2, column=2, value="=COUNTIF('Extras'!E:E,\"Sim\")")

    ws0.cell(row=r0 + 4, column=1, value="% álbum base completo")
    ws0.cell(row=r0 + 4, column=2, value="=(COUNTIF('Página Inicial'!E:E,\"Sim\")+COUNTIF('Seleções'!H:H,\"Sim\")+COUNTIF('Museu FIFA'!E:E,\"Sim\"))/980")
    ws0.cell(row=r0 + 4, column=2).number_format = "0.00%"

    ws0.cell(row=r0 + 6, column=1, value="Total de duplicatas (soma)")
    ws0.cell(row=r0 + 6, column=2, value="=SUM('Página Inicial'!F:F)+SUM('Seleções'!I:I)+SUM('Museu FIFA'!F:F)+SUM('Coca-Cola'!F:F)+SUM('Extras'!F:F)")

    autosize_columns(ws0)

    foil_type = foil_type_core

    # --- Página Inicial ---
    w1 = wb.create_sheet("Página Inicial")
    headers1 = ["#", "Código", "Descrição no álbum", "Tipo", "Tenho", "Duplicatas"]
    for c, h in enumerate(headers1, 1):
        w1.cell(row=1, column=c, value=h)
    apply_header(w1, 1, len(headers1))
    for i, (code, desc) in enumerate(intro, start=1):
        row = i + 1
        w1.cell(row=row, column=1, value=i)
        w1.cell(row=row, column=2, value=code)
        w1.cell(row=row, column=3, value=desc)
        w1.cell(row=row, column=4, value=foil_type(desc))
        w1.cell(row=row, column=6, value=0)
    add_tenho_validation(w1, "E", 2, 1 + len(intro))
    w1.freeze_panes = "A2"
    autosize_columns(w1)

    # --- Seleções (960) ---
    ws = wb.create_sheet("Seleções")
    hs = [
        "Grupo",
        "Sigla",
        "Seleção",
        "#",
        "Código",
        "Descrição no álbum",
        "Tipo",
        "Tenho",
        "Duplicatas",
    ]
    for c, h in enumerate(hs, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header(ws, 1, len(hs))
    row_n = 2
    for tm in teams_meta:
        for j, (code, desc) in enumerate(tm["block"], start=1):
            ws.cell(row=row_n, column=1, value=tm["grupo"])
            ws.cell(row=row_n, column=2, value=tm["sigla"])
            ws.cell(row=row_n, column=3, value=tm["pais"])
            ws.cell(row=row_n, column=4, value=j)
            ws.cell(row=row_n, column=5, value=code)
            ws.cell(row=row_n, column=6, value=desc)
            ws.cell(row=row_n, column=7, value=foil_type(desc))
            ws.cell(row=row_n, column=9, value=0)
            row_n += 1
    add_tenho_validation(ws, "H", 2, row_n - 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)

    # --- Museu FIFA ---
    wm = wb.create_sheet("Museu FIFA")
    hm = ["#", "Código", "Descrição no álbum", "Tipo", "Tenho", "Duplicatas"]
    for c, h in enumerate(hm, 1):
        wm.cell(row=1, column=c, value=h)
    apply_header(wm, 1, len(hm))
    for i, (code, desc) in enumerate(museum, start=1):
        wm.cell(row=i + 1, column=1, value=i)
        wm.cell(row=i + 1, column=2, value=code)
        wm.cell(row=i + 1, column=3, value=desc)
        wm.cell(row=i + 1, column=4, value=foil_type(desc))
        wm.cell(row=i + 1, column=6, value=0)
    add_tenho_validation(wm, "E", 2, 1 + len(museum))
    wm.freeze_panes = "A2"
    autosize_columns(wm)

    # --- Siglas ---
    wx = wb.create_sheet("Siglas")
    xh = ["Sigla FIFA", "Nome da seleção", "Grupo", "Figurinhas (20)", "Obs."]
    for c, h in enumerate(xh, 1):
        wx.cell(row=1, column=c, value=h)
    apply_header(wx, 1, len(xh))
    for i, tm in enumerate(teams_meta, start=2):
        wx.cell(row=i, column=1, value=tm["sigla"])
        wx.cell(row=i, column=2, value=tm["pais"])
        wx.cell(row=i, column=3, value=tm["grupo"])
        wx.cell(row=i, column=4, value="20")
        if tm["sigla_panini"] != tm["sigla"]:
            wx.cell(row=i, column=5, value=f"Códigos Panini usam «{tm['sigla_panini']}»")
    wx.freeze_panes = "A2"
    wx.auto_filter.ref = wx.dimensions
    autosize_columns(wx)

    # --- Grupos (visão mapa) ---
    wg = wb.create_sheet("Grupos")
    gh = ["Grupo", "Seleção 1", "Seleção 2", "Seleção 3", "Seleção 4"]
    for c, h in enumerate(gh, 1):
        wg.cell(row=1, column=c, value=h)
    apply_header(wg, 1, len(gh))
    for gi in range(12):
        slice_t = teams_meta[gi * 4 : (gi + 1) * 4]
        row = gi + 2
        wg.cell(row=row, column=1, value=f"Grupo {chr(ord('A') + gi)}")
        for j, tm in enumerate(slice_t):
            wg.cell(row=row, column=2 + j, value=f"{tm['pais']} ({tm['sigla']})")
    wg.freeze_panes = "A2"
    autosize_columns(wg)

    # --- Coca-Cola ---
    wc = wb.create_sheet("Coca-Cola")
    ch = ["# Promo", "Jogador", "Seleção", "Origem", "Tenho", "Duplicatas"]
    for c, h in enumerate(ch, 1):
        wc.cell(row=1, column=c, value=h)
    apply_header(wc, 1, len(ch))
    for i, (num, player, sel) in enumerate(COCA_COLA, start=2):
        wc.cell(row=i, column=1, value=num)
        wc.cell(row=i, column=2, value=player)
        wc.cell(row=i, column=3, value=sel)
        wc.cell(row=i, column=4, value="Rótulo Coca-Cola EUA / fora dos pacotes")
        wc.cell(row=i, column=6, value=0)
    add_tenho_validation(wc, "E", 2, 1 + len(COCA_COLA))
    wc.freeze_panes = "A2"
    autosize_columns(wc)

    # --- Extras ---
    we = wb.create_sheet("Extras")
    eh = ["#", "Jogador", "Seleção", "Notas", "Tenho", "Duplicatas"]
    for c, h in enumerate(eh, 1):
        we.cell(row=1, column=c, value=h)
    apply_header(we, 1, len(eh))
    for i, (player, sel) in enumerate(EXTRAS, start=2):
        we.cell(row=i, column=1, value=i - 1)
        we.cell(row=i, column=2, value=player)
        we.cell(row=i, column=3, value=sel)
        we.cell(row=i, column=4, value="Extra internacional (~1:100 pacotes); sem número no álbum")
        we.cell(row=i, column=6, value=0)
    add_tenho_validation(we, "E", 2, 1 + len(EXTRAS))
    we.freeze_panes = "A2"
    autosize_columns(we)

    # --- Todas (980) ---
    wa = wb.create_sheet("Todas (980)")
    ah = ["Secção", "Grupo", "Sigla", "Seleção", "Código", "Descrição", "Tipo", "Tenho", "Dup."]
    for c, h in enumerate(ah, 1):
        wa.cell(row=1, column=c, value=h)
    apply_header(wa, 1, len(ah))
    rr = 2
    for code, desc in intro:
        wa.cell(row=rr, column=1, value="Página inicial")
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    for tm in teams_meta:
        for code, desc in tm["block"]:
            wa.cell(row=rr, column=1, value="Seleção")
            wa.cell(row=rr, column=2, value=tm["grupo"])
            wa.cell(row=rr, column=3, value=tm["sigla"])
            wa.cell(row=rr, column=4, value=tm["pais"])
            wa.cell(row=rr, column=5, value=code)
            wa.cell(row=rr, column=6, value=desc)
            wa.cell(row=rr, column=7, value=foil_type(desc))
            wa.cell(row=rr, column=9, value=0)
            rr += 1
    for code, desc in museum:
        wa.cell(row=rr, column=1, value="Museu FIFA")
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    add_tenho_validation(wa, "H", 2, rr - 1)
    wa.freeze_panes = "A2"
    wa.auto_filter.ref = wa.dimensions
    autosize_columns(wa)

    ws0.cell(row=r0 + 8, column=1, value="Faltam (base)")
    ws0.cell(row=r0 + 8, column=2, value="=980-(COUNTIF('Página Inicial'!E:E,\"Sim\")+COUNTIF('Seleções'!H:H,\"Sim\")+COUNTIF('Museu FIFA'!E:E,\"Sim\"))")

    wb.save(OUT)
    print(f"Escrito: {OUT}")


if __name__ == "__main__":
    main()
