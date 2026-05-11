#!/usr/bin/env python3
"""
Versão 3 — álbum Panini FIFA World Cup 2026: guia editorial, dashboard completo,
progresso por seleção, diário de pacotes e estética coerente (sem AutoFilter+Table).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from generate_xlsx_v2 import (
    attach_table,
    add_tenho_conditional,
    add_tenho_validation,
    autosize_columns,
)
from wc2026_album_core import CHECKLIST_DEFAULT, COCA_COLA, EXTRAS, load_album

DIR = Path(__file__).resolve().parent
OUT_V3 = DIR / "FIFA-World-Cup-2026-Panini-Controle-v3.xlsx"

# --- Tema «night pitch + gold» ---
THEME = {
    "night": "0F172A",
    "panel": "1E293B",
    "header": "0F2744",
    "gold": "CA8A04",
    "mint": "34D399",
    "cream": "FEFCE8",
    "muted": "94A3B8",
    "band": "F1F5F9",
}
COL_SIM = "C6EFCE"
COL_NAO = "F8D7DA"
COL_TROCA = "FFF3CD"
THIN = Side(style="thin", color="334155")


def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def foil_type(desc: str) -> str:
    d = desc.strip()
    return "FOIL" if d.endswith("FOIL") else "Base"


def apply_header_v3(ws, row: int, cols: int) -> None:
    gold_b = Border(
        left=THIN,
        right=THIN,
        top=THIN,
        bottom=Side(style="thick", color=THEME["gold"]),
    )
    fill = PatternFill("solid", fgColor=THEME["header"])
    font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = gold_b
    ws.row_dimensions[row].height = 26


def tab(ws, color_hex: str) -> None:
    ws.sheet_properties.tabColor = color_hex.replace("#", "")


def band_grupos_selecoes(ws, last_row: int) -> None:
    """Faixas alternadas a cada 80 linhas (4 seleções × 20 figurinhas)."""
    rng = f"A2:{get_column_letter(ws.max_column)}{last_row}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=["=MOD(INT((ROW()-2)/80),2)=0"],
            fill=PatternFill("solid", fgColor=THEME["band"]),
        ),
    )


def hl(sheet_name: str, cell: str, label: str) -> str:
    """Fórmula HYPERLINK interna (Excel)."""
    safe = sheet_name.replace("'", "''")
    return f"=HYPERLINK(\"#'{safe}'!{cell}\",\"{label}\")"


def build_guide(wb: Workbook) -> None:
    wg = wb["Guia"]
    wg.sheet_view.showGridLines = False
    for r in range(1, 36):
        for c in range(1, 10):
            wg.cell(row=r, column=c).fill = PatternFill("solid", fgColor=THEME["night"])

    wg.merge_cells("A2:H4")
    t = wg["A2"]
    t.value = "Panini · FIFA World Cup 2026"
    t.font = Font(name="Calibri", size=28, bold=True, color=THEME["cream"])
    t.alignment = Alignment(horizontal="left", vertical="center")

    wg.merge_cells("A5:H6")
    wg["A5"] = (
        "Controlo completo da coleção — marque «Tenho», duplicatas e use o Dashboard ao vivo. "
        "Versão 3: mapa do livro, progresso por seleção e espaço para registar pacotes."
    )
    wg["A5"].font = Font(size=11, color=THEME["muted"])
    wg["A5"].alignment = Alignment(wrap_text=True, vertical="center")

    wg["A8"] = "Mapa do livro"
    wg["A8"].font = Font(size=14, bold=True, color=THEME["cream"])
    wg.merge_cells("A9:H9")
    wg["A9"].fill = PatternFill("solid", fgColor=THEME["panel"])

    rows = [
        ("Painel principal", hl("Dashboard", "A1", "► Dashboard · KPIs e gráficos")),
        ("Introdução do álbum", hl("Página Inicial", "A1", "► Página inicial · 9 figurinhas")),
        ("48 seleções", hl("Seleções", "A1", "► Seleções · lista completa")),
        ("Museu dos campeões", hl("Museu FIFA", "A1", "► Museu FIFA")),
        ("Promo Coca-Cola", hl("Coca-Cola", "A1", "► Coca-Cola · 12")),
        ("Extras internacionais", hl("Extras", "A1", "► Extras · 20")),
        ("Resumo por país", hl("Por seleção", "A1", "► Por seleção · 48 linhas")),
        ("Diário de pacotes", hl("Pacotes", "A1", "► Pacotes · registo livre")),
        ("Motor & grupos", hl("Estatísticas", "A1", "► Estatísticas")),
        ("Siglas & referência", hl("Siglas", "A1", "► Siglas")),
        ("Mapa dos grupos", hl("Grupos", "A1", "► Grupos A–L")),
        ("Lista única 980", hl("Todas (980)", "A1", "► Todas (980)")),
    ]
    r = 10
    wg.cell(row=r, column=1, value="Secção")
    wg.cell(row=r, column=2, value="Abrir")
    for c in (1, 2):
        wg.cell(row=r, column=c).font = Font(bold=True, color=THEME["cream"])
        wg.cell(row=r, column=c).fill = PatternFill("solid", fgColor=THEME["panel"])
    r += 1
    for title, formula in rows:
        wg.cell(row=r, column=1, value=title)
        wg.cell(row=r, column=2, value=formula)
        wg.cell(row=r, column=1).font = Font(color=THEME["muted"])
        wg.cell(row=r, column=2).font = Font(color=THEME["mint"], underline="single")
        r += 1

    r += 1
    wg.cell(row=r, column=1, value="Legenda · coluna «Tenho»")
    wg.cell(row=r, column=1).font = Font(bold=True, color=THEME["cream"])
    r += 1
    legend = [
        ("Sim", "Já colou no álbum"),
        ("Não", "Ainda falta"),
        ("Falta trocar", "Tem repetida — precisa de troca"),
    ]
    for lab, desc in legend:
        wg.cell(row=r, column=1, value=lab)
        wg.cell(row=r, column=2, value=desc)
        wg.cell(row=r, column=1).fill = PatternFill(
            "solid",
            fgColor={"Sim": COL_SIM, "Não": COL_NAO, "Falta trocar": COL_TROCA}[lab].replace("#", ""),
        )
        r += 1

    wg.column_dimensions["A"].width = 28
    wg.column_dimensions["B"].width = 52
    tab(wg, THEME["gold"])


def build_data_sheets_v3(wb: Workbook, intro, museum, teams_meta) -> None:
    # Página Inicial
    w1 = wb.create_sheet("Página Inicial")
    tab(w1, "2563EB")
    h1 = ["#", "Código", "Descrição no álbum", "Tipo", "Tenho", "Duplicatas"]
    for c, h in enumerate(h1, 1):
        w1.cell(row=1, column=c, value=h)
    apply_header_v3(w1, 1, len(h1))
    for i, (code, desc) in enumerate(intro, start=1):
        row = i + 1
        w1.cell(row=row, column=1, value=i)
        w1.cell(row=row, column=2, value=code)
        w1.cell(row=row, column=3, value=desc)
        w1.cell(row=row, column=4, value=foil_type(desc))
        w1.cell(row=row, column=6, value=0)
        w1.row_dimensions[row].height = 22
    last_pi = 1 + len(intro)
    add_tenho_validation(w1, "E", 2, last_pi)
    add_tenho_conditional(w1, "E", 2, last_pi)
    w1.column_dimensions["E"].width = 16
    w1.column_dimensions["C"].width = 44
    w1.freeze_panes = "E2"
    attach_table(w1, f"A1:F{last_pi}", "tblPaginaInicial", stripe=True)

    # Seleções
    ws = wb.create_sheet("Seleções")
    tab(ws, "059669")
    hs = ["Grupo", "Sigla", "Seleção", "#", "Código", "Descrição no álbum", "Tipo", "Tenho", "Duplicatas"]
    for c, h in enumerate(hs, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header_v3(ws, 1, len(hs))
    row_n = 2
    for tm in teams_meta:
        for j, (code, desc) in enumerate(tm["block"], start=1):
            ws.cell(row=row_n, column=1, value=tm["grupo"])
            ws.cell(row=row_n, column=2, value=tm["sigla"])
            ws.cell(row=row_n, column=3, value=tm["pais"])
            ws.cell(row=row_n, column=4, value=j)
            ws.cell(row=row_n, column=5, value=code)
            ws.cell(row=row_n, column=6, value=desc)
            ws.cell(row=row_n, column=7, value=foil_type(desc))
            ws.cell(row=row_n, column=9, value=0)
            ws.row_dimensions[row_n].height = 20
            row_n += 1
    last_sel = row_n - 1
    add_tenho_validation(ws, "H", 2, last_sel)
    add_tenho_conditional(ws, "H", 2, last_sel)
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["F"].width = 38
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "E2"
    attach_table(ws, f"A1:I{last_sel}", "tblSelecoes", stripe=True)
    band_grupos_selecoes(ws, last_sel)

    # Museu
    wm = wb.create_sheet("Museu FIFA")
    tab(wm, "7C3AED")
    hm = ["#", "Código", "Descrição no álbum", "Tipo", "Tenho", "Duplicatas"]
    for c, h in enumerate(hm, 1):
        wm.cell(row=1, column=c, value=h)
    apply_header_v3(wm, 1, len(hm))
    for i, (code, desc) in enumerate(museum, start=1):
        wm.cell(row=i + 1, column=1, value=i)
        wm.cell(row=i + 1, column=2, value=code)
        wm.cell(row=i + 1, column=3, value=desc)
        wm.cell(row=i + 1, column=4, value=foil_type(desc))
        wm.cell(row=i + 1, column=6, value=0)
        wm.row_dimensions[i + 1].height = 22
    last_mu = 1 + len(museum)
    add_tenho_validation(wm, "E", 2, last_mu)
    add_tenho_conditional(wm, "E", 2, last_mu)
    wm.column_dimensions["E"].width = 16
    wm.column_dimensions["C"].width = 42
    wm.freeze_panes = "E2"
    attach_table(wm, f"A1:F{last_mu}", "tblMuseu", stripe=True)

    # Coca-Cola
    wc = wb.create_sheet("Coca-Cola")
    tab(wc, "DC2626")
    ch = ["# Promo", "Jogador", "Seleção", "Origem", "Tenho", "Duplicatas"]
    for c, h in enumerate(ch, 1):
        wc.cell(row=1, column=c, value=h)
    apply_header_v3(wc, 1, len(ch))
    for i, (num, player, sel) in enumerate(COCA_COLA, start=2):
        wc.cell(row=i, column=1, value=num)
        wc.cell(row=i, column=2, value=player)
        wc.cell(row=i, column=3, value=sel)
        wc.cell(row=i, column=4, value="Rótulo Coca-Cola EUA")
        wc.cell(row=i, column=6, value=0)
        wc.row_dimensions[i].height = 22
    last_cc = 1 + len(COCA_COLA)
    add_tenho_validation(wc, "E", 2, last_cc)
    add_tenho_conditional(wc, "E", 2, last_cc)
    wc.column_dimensions["E"].width = 16
    wc.freeze_panes = "E2"
    attach_table(wc, f"A1:F{last_cc}", "tblCocaCola", stripe=True)

    # Extras
    we = wb.create_sheet("Extras")
    tab(we, "EA580C")
    eh = ["#", "Jogador", "Seleção", "Notas", "Tenho", "Duplicatas"]
    for c, h in enumerate(eh, 1):
        we.cell(row=1, column=c, value=h)
    apply_header_v3(we, 1, len(eh))
    for i, (player, sel) in enumerate(EXTRAS, start=2):
        we.cell(row=i, column=1, value=i - 1)
        we.cell(row=i, column=2, value=player)
        we.cell(row=i, column=3, value=sel)
        we.cell(row=i, column=4, value="Extra internacional (sem nº no álbum)")
        we.cell(row=i, column=6, value=0)
        we.row_dimensions[i].height = 22
    last_ex = 1 + len(EXTRAS)
    add_tenho_validation(we, "E", 2, last_ex)
    add_tenho_conditional(we, "E", 2, last_ex)
    we.column_dimensions["E"].width = 16
    we.freeze_panes = "E2"
    attach_table(we, f"A1:F{last_ex}", "tblExtras", stripe=True)

    # Siglas
    wx = wb.create_sheet("Siglas")
    tab(wx, "64748B")
    xh = ["Sigla FIFA", "Nome da seleção", "Grupo", "Figurinhas", "Obs."]
    for c, h in enumerate(xh, 1):
        wx.cell(row=1, column=c, value=h)
    apply_header_v3(wx, 1, len(xh))
    for i, tm in enumerate(teams_meta, start=2):
        wx.cell(row=i, column=1, value=tm["sigla"])
        wx.cell(row=i, column=2, value=tm["pais"])
        wx.cell(row=i, column=3, value=tm["grupo"])
        wx.cell(row=i, column=4, value=20)
        if tm["sigla_panini"] != tm["sigla"]:
            wx.cell(row=i, column=5, value=f"Panini: «{tm['sigla_panini']}»")
    wx.freeze_panes = "A2"
    wx.auto_filter.ref = wx.dimensions
    autosize_columns(wx)

    # Grupos
    wg = wb.create_sheet("Grupos")
    tab(wg, "475569")
    gh = ["Grupo", "Seleção 1", "Seleção 2", "Seleção 3", "Seleção 4"]
    for c, h in enumerate(gh, 1):
        wg.cell(row=1, column=c, value=h)
    apply_header_v3(wg, 1, len(gh))
    for gi in range(12):
        slice_t = teams_meta[gi * 4 : (gi + 1) * 4]
        row = gi + 2
        wg.cell(row=row, column=1, value=f"Grupo {chr(ord('A') + gi)}")
        for j, tm in enumerate(slice_t):
            wg.cell(row=row, column=2 + j, value=f"{tm['pais']} ({tm['sigla']})")
    wg.freeze_panes = "A2"
    autosize_columns(wg)

    # Todas (980)
    wa = wb.create_sheet("Todas (980)")
    tab(wa, "0369A1")
    ah = ["Secção", "Grupo", "Sigla", "Seleção", "Código", "Descrição", "Tipo", "Tenho", "Dup."]
    for c, h in enumerate(ah, 1):
        wa.cell(row=1, column=c, value=h)
    apply_header_v3(wa, 1, len(ah))
    rr = 2
    for code, desc in intro:
        wa.cell(row=rr, column=1, value="Página inicial")
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    for tm in teams_meta:
        for code, desc in tm["block"]:
            wa.cell(row=rr, column=1, value="Seleção")
            wa.cell(row=rr, column=2, value=tm["grupo"])
            wa.cell(row=rr, column=3, value=tm["sigla"])
            wa.cell(row=rr, column=4, value=tm["pais"])
            wa.cell(row=rr, column=5, value=code)
            wa.cell(row=rr, column=6, value=desc)
            wa.cell(row=rr, column=7, value=foil_type(desc))
            wa.cell(row=rr, column=9, value=0)
            rr += 1
    for code, desc in museum:
        wa.cell(row=rr, column=1, value="Museu FIFA")
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    add_tenho_validation(wa, "H", 2, rr - 1)
    add_tenho_conditional(wa, "H", 2, rr - 1)
    wa.freeze_panes = "E2"
    wa.column_dimensions["H"].width = 16
    attach_table(wa, f"A1:I{rr - 1}", "tblTodas980", stripe=True)
    autosize_columns(wa)


def build_por_selecao(wb: Workbook, teams_meta) -> None:
    wp = wb.create_sheet("Por seleção")
    tab(wp, THEME["mint"])
    heads = [
        "#",
        "Seleção",
        "Sigla",
        "Grupo",
        "Meta",
        "Coladas",
        "Pct",
        "Progresso",
        "Faltam",
    ]
    for c, h in enumerate(heads, 1):
        wp.cell(row=1, column=c, value=h)
    apply_header_v3(wp, 1, len(heads))
    for i, tm in enumerate(teams_meta, start=2):
        wp.cell(row=i, column=1, value=i - 1)
        wp.cell(row=i, column=2, value=tm["pais"])
        wp.cell(row=i, column=3, value=tm["sigla"])
        wp.cell(row=i, column=4, value=tm["grupo"])
        wp.cell(row=i, column=5, value=20)
        wp.cell(
            row=i,
            column=6,
            value=f'=COUNTIFS(tblSelecoes[Sigla],C{i},tblSelecoes[Tenho],"Sim")',
        )
        wp.cell(row=i, column=7, value=f"=IF(E{i}>0,F{i}/E{i},0)")
        wp.cell(
            row=i,
            column=8,
            value=f'=REPT("█",ROUND(G{i}*14,0))&REPT("░",14-ROUND(G{i}*14,0))',
        )
        wp.cell(row=i, column=9, value=f"=E{i}-F{i}")
        wp.cell(row=i, column=7).number_format = "0%"
        wp.row_dimensions[i].height = 21
    attach_table(wp, f"A1:I{1 + len(teams_meta)}", "tblPorSelecao", stripe=True)
    wp.freeze_panes = "F2"
    wp.column_dimensions["B"].width = 26
    wp.column_dimensions["C"].width = 8
    wp.column_dimensions["D"].width = 11
    wp.column_dimensions["H"].width = 18
    autosize_columns(wp, max_width=40)


def build_pacotes(wb: Workbook) -> None:
    wp = wb.create_sheet("Pacotes")
    tab(wp, "A855F7")
    heads = ["Data", "Pacotes abertos", "Figurinhas novas (estim.)", "Notas"]
    for c, h in enumerate(heads, 1):
        wp.cell(row=1, column=c, value=h)
    apply_header_v3(wp, 1, len(heads))
    for r in range(2, 42):
        wp.cell(row=r, column=1, value=None)
        wp.row_dimensions[r].height = 20
    attach_table(wp, "A1:D41", "tblPacotes", stripe=True)
    wp.freeze_panes = "A2"
    wp.column_dimensions["A"].width = 14
    wp.column_dimensions["B"].width = 16
    wp.column_dimensions["C"].width = 22
    wp.column_dimensions["D"].width = 44


def build_estatisticas_v3(wb: Workbook) -> dict:
    """Motor numérico + grupos + ranking (reusa estrutura da v2)."""
    ws = wb.create_sheet("Estatísticas")
    tab(ws, THEME["mint"])

    ws.merge_cells("A1:J1")
    ws["A1"] = "Panini WC 2026 · Motor analítico"
    ws["A1"].font = Font(size=15, bold=True, color=THEME["header"])
    ws["A1"].alignment = Alignment(horizontal="left")

    ws["A3"] = "Secção"
    ws["B3"] = "Meta"
    ws["C3"] = "Com «Sim»"
    ws["D3"] = "%"
    ws["E3"] = "Faltam"
    for c in range(1, 6):
        ws.cell(row=3, column=c).font = Font(bold=True)
        ws.cell(row=3, column=c).fill = PatternFill("solid", fgColor="E8EEF7")

    rows_sec = [
        ("Página inicial", 9, 'COUNTIF(tblPaginaInicial[Tenho],"Sim")'),
        ("Seleções (48×20)", 960, 'COUNTIF(tblSelecoes[Tenho],"Sim")'),
        ("Museu FIFA", 11, 'COUNTIF(tblMuseu[Tenho],"Sim")'),
    ]
    r = 4
    for label, meta_val, formula_sim in rows_sec:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=meta_val)
        ws.cell(row=r, column=3, value=f"={formula_sim}")
        ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
        ws.cell(row=r, column=5, value=f"=B{r}-C{r}")
        ws.cell(row=r, column=4).number_format = "0.0%"
        r += 1

    ws.cell(row=r, column=1, value="ÁLBUM BASE (total)")
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=2, value=980)
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(
        row=r,
        column=3,
        value=(
            "=COUNTIF(tblPaginaInicial[Tenho],\"Sim\")"
            "+COUNTIF(tblSelecoes[Tenho],\"Sim\")"
            "+COUNTIF(tblMuseu[Tenho],\"Sim\")"
        ),
    )
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=5, value=f"=B{r}-C{r}")
    ws.cell(row=r, column=4).number_format = "0.00%"
    row_album_total = r
    row_sec_first = 4
    row_sec_last = 6

    r += 2
    ws.cell(row=r, column=1, value="Promo & extras")
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value="Item")
    ws.cell(row=r, column=2, value="Meta")
    ws.cell(row=r, column=3, value="Com «Sim»")
    ws.cell(row=r, column=4, value="%")
    for c in range(1, 5):
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Coca-Cola")
    ws.cell(row=r, column=2, value=len(COCA_COLA))
    ws.cell(row=r, column=3, value='=COUNTIF(tblCocaCola[Tenho],"Sim")')
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=4).number_format = "0%"
    r += 1
    ws.cell(row=r, column=1, value="Extras internacionais")
    ws.cell(row=r, column=2, value=len(EXTRAS))
    ws.cell(row=r, column=3, value='=COUNTIF(tblExtras[Tenho],"Sim")')
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=4).number_format = "0%"
    row_ext_end = r

    r = row_ext_end + 3
    ws.cell(row=r, column=1, value="FOIL vs Base (com «Sim»)")
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value="FOIL")
    ws.cell(
        row=r,
        column=2,
        value=(
            "=SUMPRODUCT((tblPaginaInicial[Tipo]=\"FOIL\")*(tblPaginaInicial[Tenho]=\"Sim\"))"
            "+SUMPRODUCT((tblSelecoes[Tipo]=\"FOIL\")*(tblSelecoes[Tenho]=\"Sim\"))"
            "+SUMPRODUCT((tblMuseu[Tipo]=\"FOIL\")*(tblMuseu[Tenho]=\"Sim\"))"
        ),
    )
    r += 1
    ws.cell(row=r, column=1, value="Base")
    ws.cell(
        row=r,
        column=2,
        value=(
            "=SUMPRODUCT((tblPaginaInicial[Tipo]=\"Base\")*(tblPaginaInicial[Tenho]=\"Sim\"))"
            "+SUMPRODUCT((tblSelecoes[Tipo]=\"Base\")*(tblSelecoes[Tenho]=\"Sim\"))"
            "+SUMPRODUCT((tblMuseu[Tipo]=\"Base\")*(tblMuseu[Tenho]=\"Sim\"))"
        ),
    )
    row_foil_end = r

    r = row_foil_end + 3
    ws.cell(row=r, column=1, value="Progresso por grupo")
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    grp_header_row = r
    ws.cell(row=r, column=1, value="Grupo")
    ws.cell(row=r, column=2, value="Meta")
    ws.cell(row=r, column=3, value="Com Sim")
    ws.cell(row=r, column=4, value="%")
    ws.cell(row=r, column=5, value="Faltam")
    ws.cell(row=r, column=6, value="Barra")
    ws.cell(row=r, column=7, value="Rank")
    for c in range(1, 8):
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="D6E4FF")
    r += 1
    grp_first_row = r

    for gi in range(12):
        gname = f"Grupo {chr(ord('A') + gi)}"
        row = r + gi
        ws.cell(row=row, column=1, value=gname)
        ws.cell(row=row, column=2, value=80)
        ws.cell(
            row=row,
            column=3,
            value=f'=COUNTIFS(tblSelecoes[Tenho],"Sim",tblSelecoes[Grupo],"{gname}")',
        )
        ws.cell(row=row, column=4, value=f"=IF(B{row}>0,C{row}/B{row},0)")
        ws.cell(row=row, column=5, value=f"=B{row}-C{row}")
        ws.cell(row=row, column=4).number_format = "0%"
        ws.cell(
            row=row,
            column=6,
            value=f'=REPT("█",ROUND(D{row}*12,0))&REPT("·",12-ROUND(D{row}*12,0))',
        )
        ws.cell(
            row=row,
            column=7,
            value=f"=RANK.EQ(D{row},D${grp_first_row}:D${grp_first_row + 11},0)",
        )
    grp_last_row = r + 11
    attach_table(ws, f"A{grp_header_row}:G{grp_last_row}", "tblPorGrupo", stripe=True)

    r = grp_last_row + 3
    ws.cell(row=r, column=1, value="Soma duplicatas")
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(
        row=r,
        column=2,
        value=(
            "=SUM(tblPaginaInicial[Duplicatas])+SUM(tblSelecoes[Duplicatas])"
            "+SUM(tblMuseu[Duplicatas])+SUM(tblCocaCola[Duplicatas])+SUM(tblExtras[Duplicatas])"
        ),
    )

    autosize_columns(ws)

    return {
        "row_album_total": row_album_total,
        "grp_first_row": grp_first_row,
        "grp_last_row": grp_last_row,
        "row_sec_first": row_sec_first,
        "row_sec_last": row_sec_last,
    }


def build_dashboard_v3(wb: Workbook, stats: dict) -> None:
    wd = wb["Dashboard"]
    tab(wd, THEME["gold"])
    rt = stats["row_album_total"]
    gf = stats["grp_first_row"]
    gl = stats["grp_last_row"]
    rsf = stats["row_sec_first"]
    rsl = stats["row_sec_last"]

    for row in range(1, 48):
        for col in range(1, 24):
            wd.cell(row=row, column=col).fill = PatternFill("solid", fgColor=THEME["night"])
    wd.sheet_view.showGridLines = False

    wd.merge_cells("A1:W2")
    wd["A1"] = "Painel da coleção"
    wd["A1"].font = Font(name="Calibri", size=24, bold=True, color=THEME["cream"])
    wd["A1"].alignment = Alignment(horizontal="left", vertical="center")

    wd.merge_cells("A3:W3")
    wd["A3"] = (
        "Resumo dinâmico · as figurinhas base ligam-se às folhas «Página Inicial», «Seleções» e «Museu FIFA». "
        "Consulte «Guia» para navegar."
    )
    wd["A3"].font = Font(size=10, color=THEME["muted"])
    wd["A3"].alignment = Alignment(wrap_text=True)

    cards = [
        ("B5:D5", "B6:D6", "Álbum base", f"=Estatísticas!C{rt}&\" / \"&Estatísticas!B{rt}"),
        ("E5:G5", "E6:G6", "Faltam", f"=Estatísticas!E{rt}"),
        ("H5:J5", "H6:J6", "% Completo", f"=Estatísticas!D{rt}"),
        ("K5:M5", "K6:M6", "Coca-Cola", '=COUNTIF(tblCocaCola[Tenho],"Sim")&" / "&COUNTA(tblCocaCola[Tenho])'),
        ("N5:P5", "N6:P6", "Extras", '=COUNTIF(tblExtras[Tenho],"Sim")&" / "&COUNTA(tblExtras[Tenho])'),
        ("Q5:S5", "Q6:S6", "Duplicatas", "=SUM(tblPaginaInicial[Duplicatas])+SUM(tblSelecoes[Duplicatas])+SUM(tblMuseu[Duplicatas])+SUM(tblCocaCola[Duplicatas])+SUM(tblExtras[Duplicatas])"),
        ("B8:D8", "B9:D9", "Falta trocar", "=COUNTIF(tblPaginaInicial[Tenho],\"Falta trocar\")+COUNTIF(tblSelecoes[Tenho],\"Falta trocar\")+COUNTIF(tblMuseu[Tenho],\"Falta trocar\")"),
        ("E8:G8", "E9:G9", "FOIL coladas", "=SUMPRODUCT((tblPaginaInicial[Tipo]=\"FOIL\")*(tblPaginaInicial[Tenho]=\"Sim\"))+SUMPRODUCT((tblSelecoes[Tipo]=\"FOIL\")*(tblSelecoes[Tenho]=\"Sim\"))+SUMPRODUCT((tblMuseu[Tipo]=\"FOIL\")*(tblMuseu[Tenho]=\"Sim\"))"),
        ("H8:J8", "H9:J9", "Seleções completas", "=COUNTIF(tblPorSelecao[Faltam],0)"),
        ("K8:M8", "K9:M9", "Hoje", "=TODAY()"),
    ]

    for rng_title, rng_val, title, formula in cards:
        wd.merge_cells(rng_title)
        wd.merge_cells(rng_val)
        tl = rng_title.split(":")[0]
        vl = rng_val.split(":")[0]
        wd[tl].value = title
        wd[tl].font = Font(size=10, color=THEME["muted"])
        wd[tl].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cv = wd[vl]
        cv.value = formula
        cv.alignment = Alignment(horizontal="center", vertical="center")
        if "% Completo" in title:
            cv.font = Font(size=26, bold=True, color=THEME["mint"])
            cv.number_format = "0.0%"
        elif title == "Hoje":
            cv.font = Font(size=13, bold=True, color=THEME["cream"])
            cv.number_format = "dddd dd/mm/yyyy"
        else:
            cv.font = Font(size=17, bold=True, color=THEME["cream"])

    wd["U40"] = "Coladas"
    wd["V40"] = f"=Estatísticas!C{rt}"
    wd["U41"] = "Faltam"
    wd["V41"] = f"=Estatísticas!E{rt}"

    dough = DoughnutChart()
    dough.title = "Álbum base"
    dough.style = 26
    dough.add_data(Reference(wd, min_col=22, min_row=40, max_row=41), titles_from_data=False)
    dough.set_categories(Reference(wd, min_col=21, min_row=40, max_row=41))
    dough.dataLabels = DataLabelList()
    dough.dataLabels.showPercent = True
    wd.add_chart(dough, "B12")

    colc = BarChart()
    colc.type = "col"
    colc.style = 11
    colc.title = "Coladas por secção"
    colc.y_axis.title = "Figurinhas"
    colc.add_data(
        Reference(wb["Estatísticas"], min_col=3, min_row=rsf, max_row=rsl),
        titles_from_data=False,
    )
    colc.set_categories(Reference(wb["Estatísticas"], min_col=1, min_row=rsf, max_row=rsl))
    wd.add_chart(colc, "B28")

    bars = BarChart()
    bars.type = "bar"
    bars.style = 12
    bars.title = "% por grupo"
    bars.x_axis.title = "%"
    bars.add_data(Reference(wb["Estatísticas"], min_col=4, min_row=gf, max_row=gl), titles_from_data=False)
    bars.set_categories(Reference(wb["Estatísticas"], min_col=1, min_row=gf, max_row=gl))
    wd.add_chart(bars, "M12")

    wd.merge_cells("B42:W44")
    wd["B42"] = (
        "Dica: «Por seleção» mostra cada país numa linha com barra de progresso. "
        "«Pacotes» serve para anotar datas e volumes — não entra nas fórmulas do álbum."
    )
    wd["B42"].font = Font(size=10, color=THEME["muted"])
    wd["B42"].alignment = Alignment(wrap_text=True, vertical="top")

    for rr in (40, 41):
        wd.row_dimensions[rr].hidden = True

    wd.column_dimensions["B"].width = 12


def main() -> None:
    intro, museum, teams_meta = load_album(CHECKLIST_DEFAULT)
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Guia", 0)
    wb.create_sheet("Dashboard", 1)

    build_guide(wb)
    build_data_sheets_v3(wb, intro, museum, teams_meta)
    build_por_selecao(wb, teams_meta)
    build_pacotes(wb)
    stats = build_estatisticas_v3(wb)
    build_dashboard_v3(wb, stats)

    wb.save(OUT_V3)
    print(f"Escrito: {OUT_V3}")


if __name__ == "__main__":
    main()
