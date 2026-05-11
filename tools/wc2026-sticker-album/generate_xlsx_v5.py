#!/usr/bin/env python3
"""
Versão 5 — UM único Excel; idioma dinâmico via folha Config (PT / EN).

Na folha «Config» escolha PT ou EN em B2. Listas, Guia e cartões do Dashboard
actualizam via fórmulas. Colunas de dados usam nomes neutros (Status, Dup, …)
e as contagens aceitam valores em português OU inglês (troca de idioma não apaga dados).

  python3 generate_xlsx_v5.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from generate_xlsx_v2 import attach_table, autosize_columns
from wc2026_album_core import CHECKLIST_DEFAULT, COCA_COLA, EXTRAS, load_album

DIR = Path(__file__).resolve().parent
OUT_V5 = DIR / "FIFA-World-Cup-2026-Panini-Controle-v5.xlsx"

THEME = {
    "night": "0F172A",
    "panel": "1E293B",
    "mat": "151F32",
    "header": "0F2744",
    "gold": "CA8A04",
    "gold_bar": "D4A012",
    "mint": "34D399",
    "cream": "FEFCE8",
    "muted": "94A3B8",
    "band": "F1F5F9",
}
COL_SIM = "C6EFCE"
COL_NAO = "F8D7DA"
COL_TROCA = "FFF3CD"
THIN = Side(style="thin", color="334155")

# Cabeçalhos de tabela fixos (idioma-neutros) — fórmulas referenciam estes nomes.
H = {
    "idx": "Idx",
    "code": "Code",
    "sticker": "Sticker",
    "typ": "Type",
    "status": "Status",
    "dup": "Dup",
    "grp": "Group",
    "abbr": "Abbr",
    "nat": "Nation",
    "promo": "Promo",
    "player": "Player",
    "src": "Source",
    "notes": "Notes",
    "sec": "Section",
    "goal": "Goal",
    "got": "Got",
    "pct": "Pct",
    "prog": "Prog",
    "left": "Left",
    "miss": "Missing",
    "bar": "Bar",
    "date": "Date",
    "packs": "PacksOpen",
    "new_est": "NewEst",
    "fifa": "FIFA",
    "country": "Country",
    "stk": "Stickers",
    "obs": "Obs",
    "t1": "T1",
    "t2": "T2",
    "t3": "T3",
    "t4": "T4",
}


def esc(s: str) -> str:
    return s.replace('"', '""')


def ch(pt: str, en: str) -> str:
    """Fórmula CHOOSE ligada a Config!B2 ∈ {PT, EN}."""
    return f'=CHOOSE(MATCH(Config!$B$2,{{"PT";"EN"}},0),"{esc(pt)}","{esc(en)}")'


def ch_sub(pt: str, en: str) -> str:
    """Mesmo que ch(), sem '=' — para embutir em HYPERLINK(..., ...) ou outras fórmulas."""
    return f'CHOOSE(MATCH(Config!$B$2,{{"PT";"EN"}},0),"{esc(pt)}","{esc(en)}")'


def sq(name: str, cell: str) -> str:
    safe = name.replace("'", "''")
    return f"'{safe}'!{cell}"


def cnt_yes(tbl: str) -> str:
    """Conta «tem» em PT ou EN."""
    return (
        f'COUNTIF({tbl}[{H["status"]}],"Sim")+COUNTIF({tbl}[{H["status"]}],"Yes")'
    )


def cnt_no(tbl: str) -> str:
    return f'COUNTIF({tbl}[{H["status"]}],"Não")+COUNTIF({tbl}[{H["status"]}],"No")'


def cnt_trade(tbl: str) -> str:
    return (
        f'COUNTIF({tbl}[{H["status"]}],"Falta trocar")'
        f'+COUNTIF({tbl}[{H["status"]}],"Need swap")'
    )


def foil_type(desc: str) -> str:
    d = desc.strip()
    return "FOIL" if d.endswith("FOIL") else "Base"


def apply_header(ws, row: int, cols: int) -> None:
    gold_b = Border(
        left=THIN,
        right=THIN,
        top=THIN,
        bottom=Side(style="thick", color=THEME["gold"]),
    )
    fill = PatternFill("solid", fgColor=THEME["header"])
    font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = gold_b
    ws.row_dimensions[row].height = 28


def tab(ws, color_hex: str) -> None:
    ws.sheet_properties.tabColor = color_hex.replace("#", "")


def status_validation(ws, col_letter: str, sr: int, er: int) -> None:
    """Lista dinâmica: Config!D1:D3 muda com o idioma."""
    dv = DataValidation(type="list", formula1="Config!$D$1:$D$3", allow_blank=True)
    dv.errorTitle = "Lista"
    dv.error = "Escolha um valor da lista em Config." 
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{sr}:{col_letter}{er}")


def status_conditional(ws, col_letter: str, sr: int, er: int) -> None:
    """Aceita Sim/Yes, Não/No, Falta trocar/Need swap."""
    r = f"{col_letter}{sr}:{col_letter}{er}"
    col = f"${col_letter}{sr}"
    ws.conditional_formatting.add(
        r,
        FormulaRule(
            formula=[f'OR({col}="Sim",{col}="Yes")'],
            fill=PatternFill("solid", fgColor=COL_SIM.replace("#", "")),
        ),
    )
    ws.conditional_formatting.add(
        r,
        FormulaRule(
            formula=[f'OR({col}="Não",{col}="No")'],
            fill=PatternFill("solid", fgColor=COL_NAO.replace("#", "")),
        ),
    )
    ws.conditional_formatting.add(
        r,
        FormulaRule(
            formula=[f'OR({col}="Falta trocar",{col}="Need swap")'],
            fill=PatternFill("solid", fgColor=COL_TROCA.replace("#", "")),
        ),
    )


def band_grupos(ws, last_row: int) -> None:
    rng = f"A2:{get_column_letter(ws.max_column)}{last_row}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=["=MOD(INT((ROW()-2)/80),2)=0"],
            fill=PatternFill("solid", fgColor=THEME["band"]),
        ),
    )


def nav_link(sheet: str, label_expr: str) -> str:
    """HYPERLINK interno; folhas com espaço precisam de 'Nome'! no endereço."""
    s = sheet.replace("'", "''")
    # Segundo argumento: expressão (sem '=' inicial), ex. CHOOSE(...) ou texto entre aspas.
    return f'=HYPERLINK("#\'{s}\'!A1",{label_expr})'


def paint_dash_mat(wd, max_r: int = 44, max_c: int = 23) -> None:
    for row in range(4, max_r):
        for col in range(2, max_c):
            wd.cell(row=row, column=col).fill = PatternFill("solid", fgColor=THEME["mat"])


def kpi_gold_left(wd, top_left: str) -> None:
    c = wd[top_left]
    c.border = Border(left=Side(style="thick", color=THEME["gold"]), top=THIN, right=THIN, bottom=THIN)


def build_config(wb: Workbook) -> None:
    ws = wb["Config"]
    tab(ws, "F59E0B")
    ws["A1"] = "Idioma / Language"
    ws["A1"].font = Font(bold=True, size=12)
    # Lista de validação em células (funciona melhor em Excel/LibreOffice que "PT,EN" entre aspas).
    ws["G1"], ws["G2"] = "PT", "EN"
    ws["F1"] = ch("Valores para B2", "Values for B2")
    ws["F1"].font = Font(size=9, italic=True, color="666666")
    ws["B2"] = "PT"
    dv = DataValidation(type="list", formula1="=$G$1:$G$2", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("B2")
    ws["A2"] = ch("Interface (atalhos, ajuda, rótulos dinâmicos)", "Interface (help, shortcuts, dynamic labels)")
    ws["A2"].font = Font(color=THEME["muted"])
    ws["C1"] = ch("Textos da lista «Status» (mudam com B2)", "«Status» dropdown labels (follow B2)")
    ws["D1"] = '=IF(Config!$B$2="PT","Sim","Yes")'
    ws["D2"] = '=IF(Config!$B$2="PT","Não","No")'
    ws["D3"] = '=IF(Config!$B$2="PT","Falta trocar","Need swap")'
    ws["C4"] = ch(
        "Estas células alimentam as listas pendentes da coluna Status nas folhas de cromos.",
        "These cells feed the Status dropdown lists on sticker sheets.",
    )
    ws.merge_cells("C4:F5")
    ws["C4"].alignment = Alignment(wrap_text=True)

    sec_fill = PatternFill("solid", fgColor="FEF9C3")
    ws["A7"] = ch("▼ ONDE PREENCHER", "▼ WHERE TO ENTER DATA")
    ws["A7"].font = Font(bold=True, size=13, color="713F12")
    ws["A7"].fill = sec_fill
    ws.merge_cells("A7:F7")
    ws.row_dimensions[7].height = 22

    fill_instr = PatternFill("solid", fgColor="FFFBEB")
    instr = [
        ch(
            "Passo 1 — Só precisa de preencher a coluna «Status» e (opcional) «Dup» nas folhas de figurinhas; o resto é fixo.",
            "Step 1 — Only fill «Status» and optionally «Dup» on sticker sheets; other columns are fixed.",
        ),
        ch(
            "Passo 2 — Clique na célula «Status» e escolha na lista (os textos vêm de D1:D3 conforme PT ou EN em B2).",
            "Step 2 — Click «Status» and pick from the list (labels follow D1:D3 based on PT or EN in B2).",
        ),
        ch(
            "Passo 3 — «Dup» = quantas repetidas tens dessa cromo (número). Deixe 0 se não tiver duplicada.",
            "Step 3 — «Dup» = how many duplicates of that sticker (number). Use 0 if none.",
        ),
        ch(
            "• Página Inicial — coluna E (Status) · Museu FIFA — coluna E · Coca-Cola — coluna E · Extras — coluna E",
            "• Intro — column E (Status) · FIFA Museum — column E · Coca-Cola — column E · Extras — column E",
        ),
        ch(
            "• Seleções — coluna H · Todas (980) — coluna H",
            "• Teams — column H · Master 980 — column H",
        ),
        ch(
            "• Pacotes — folha «Pacotes» (datas / quantidade). · Siglas / Grupos — só consulta (sem Status).",
            "• Packs — «Pacotes» sheet (dates / counts). · Codes / Groups — read-only (no Status).",
        ),
        ch(
            "Nota: se mudares PT↔EN, o Excel não traduz o que já escreveste em «Status»; podes voltar a escolher na lista.",
            "Note: switching PT↔EN does not translate existing «Status» cells; re-pick from the list if you want.",
        ),
    ]
    for i, txt in enumerate(instr, start=8):
        ws.merge_cells(f"A{i}:F{i}")
        c = ws.cell(row=i, column=1, value=txt)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.fill = fill_instr
        ws.row_dimensions[i].height = 36 if i in (8, 9) else 30

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 6
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A8"


def build_guide(wb: Workbook) -> None:
    wg = wb["Guia"]
    wg.sheet_view.showGridLines = False
    wg.merge_cells("A1:K1")
    wg.row_dimensions[1].height = 11
    wg["A1"].fill = PatternFill("solid", fgColor=THEME["gold_bar"])

    for r in range(2, 56):
        for c in range(1, 12):
            wg.cell(row=r, column=c).fill = PatternFill("solid", fgColor=THEME["night"])

    wg.merge_cells("A3:K6")
    wg["A3"] = ch("Panini · FIFA World Cup 2026", "Panini · FIFA World Cup 2026")
    wg["A3"].font = Font(size=30, bold=True, color=THEME["cream"])

    wg.merge_cells("A8:K10")
    wg["A8"] = ch(
        "Comece pela folha «Config»: em B2 escolha PT ou EN (lista definida em G1:G2). "
        "Depois preencha só a coluna «Status» nas folhas de cromos — ver passos abaixo.",
        "Start on «Config»: pick PT or EN in B2 (see G1:G2). "
        "Then fill only the «Status» column on sticker sheets — see steps below.",
    )
    wg["A8"].font = Font(size=11, color=THEME["muted"])
    wg["A8"].alignment = Alignment(wrap_text=True, vertical="center")

    step_lines = [
        ch(
            "① «Config» · B2 = idioma da interface. D1:D3 mostram os três estados da lista «Status» (Sim/Não/Falta trocar ou Yes/No/Need swap).",
            "① «Config» · B2 = UI language. D1:D3 show the three «Status» states for dropdowns.",
        ),
        ch(
            "② O que editar: em cada folha de figurinhas use só «Status» e, se quiser, «Dup» (repetidas). Não altere código nem descrição.",
            "② What to edit: on each sticker sheet use only «Status» and optionally «Dup» (duplicates). Do not change codes or descriptions.",
        ),
        ch(
            "③ Onde está «Status»: Página Inicial → coluna E · Seleções → H · Museu FIFA → E · Coca-Cola → E · Extras → E · Todas (980) → H.",
            "③ Where «Status» lives: Intro → column E · Teams → H · Museum → E · Coca-Cola → E · Extras → E · Master 980 → H.",
        ),
        ch(
            "④ Pacotes anotados na folha «Pacotes». Siglas e Grupos são só referência. Dashboard e totais são automáticos.",
            "④ Log packs on «Pacotes». Codes & Groups are reference only. Dashboard and totals are automatic.",
        ),
    ]
    for i, txt in enumerate(step_lines, start=12):
        wg.merge_cells(f"A{i}:K{i}")
        c = wg.cell(row=i, column=1, value=txt)
        c.font = Font(size=10, color=THEME["cream"])
        c.alignment = Alignment(wrap_text=True, vertical="top")
        wg.row_dimensions[i].height = 44

    wg.merge_cells("A17:K17")
    wg["A17"] = ch("Mapa do livro", "Book map")
    wg["A17"].font = Font(size=15, bold=True, color=THEME["cream"])

    nav = [
        ("Dashboard", ch_sub("► Painel", "► Overview")),
        ("Página Inicial", ch_sub("► Intro (9)", "► Intro (9)")),
        ("Seleções", ch_sub("► Equipas", "► Teams")),
        ("Museu FIFA", ch_sub("► Museu FIFA", "► FIFA Museum")),
        ("Coca-Cola", ch_sub("► Coca-Cola", "► Coca-Cola")),
        ("Extras", ch_sub("► Extras", "► Extras")),
        ("Por seleção", ch_sub("► Por país", "► By nation")),
        ("Pacotes", ch_sub("► Pacotes", "► Packs")),
        ("Estatísticas", ch_sub("► Estatísticas", "► Statistics")),
        ("Siglas", ch_sub("► Siglas", "► Codes")),
        ("Grupos", ch_sub("► Grupos", "► Groups")),
        ("Todas (980)", ch_sub("► Lista 980", "► Master 980")),
        ("Config", ch_sub("► Idioma", "► Language")),
    ]

    r = 19
    wg.cell(row=r, column=1, value=ch("Secção", "Section"))
    wg.cell(row=r, column=2, value=ch("Atalho", "Jump"))
    for c in (1, 2):
        wg.cell(row=r, column=c).font = Font(bold=True, color=THEME["cream"])
        wg.cell(row=r, column=c).fill = PatternFill("solid", fgColor=THEME["panel"])
    r += 1
    for sh, lbl_formula in nav:
        wg.cell(row=r, column=1, value=sh)
        wg.cell(row=r, column=2, value=nav_link(sh, lbl_formula))
        wg.cell(row=r, column=1).font = Font(color=THEME["muted"])
        wg.cell(row=r, column=2).font = Font(color=THEME["mint"], underline="single")
        r += 1

    r += 1
    wg.cell(row=r, column=1, value=ch("Legenda · Status", "Legend · Status"))
    wg.cell(row=r, column=1).font = Font(bold=True, color=THEME["cream"])
    r += 1
    wg.cell(row=r, column=1, value=ch("Ver também colunas em inglês nos cabeçalhos.", "Headers use English keys for stable formulas."))
    wg.merge_cells(f"A{r}:K{r}")
    r += 1

    wg.column_dimensions["A"].width = 28
    wg.column_dimensions["B"].width = 36
    tab(wg, THEME["gold"])


def build_data(wb: Workbook, intro, museum, teams_meta) -> None:
    w1 = wb.create_sheet("Página Inicial")
    tab(w1, "2563EB")
    heads = [H["idx"], H["code"], H["sticker"], H["typ"], H["status"], H["dup"]]
    for c, h in enumerate(heads, 1):
        w1.cell(row=1, column=c, value=h)
    apply_header(w1, 1, len(heads))
    for i, (code, desc) in enumerate(intro, start=1):
        row = i + 1
        w1.cell(row=row, column=1, value=i)
        w1.cell(row=row, column=2, value=code)
        w1.cell(row=row, column=3, value=desc)
        w1.cell(row=row, column=4, value=foil_type(desc))
        w1.cell(row=row, column=6, value=0)
        w1.row_dimensions[row].height = 22
    last_pi = 1 + len(intro)
    status_validation(w1, "E", 2, last_pi)
    status_conditional(w1, "E", 2, last_pi)
    w1.column_dimensions["E"].width = 14
    w1.column_dimensions["C"].width = 44
    w1.freeze_panes = "E2"
    attach_table(w1, f"A1:F{last_pi}", "tblPaginaInicial", stripe=True)

    ws = wb.create_sheet("Seleções")
    tab(ws, "059669")
    heads = [
        H["grp"],
        H["abbr"],
        H["nat"],
        H["idx"],
        H["code"],
        H["sticker"],
        H["typ"],
        H["status"],
        H["dup"],
    ]
    for c, h in enumerate(heads, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header(ws, 1, len(heads))
    row_n = 2
    for tm in teams_meta:
        for j, (code, desc) in enumerate(tm["block"], start=1):
            ws.cell(row=row_n, column=1, value=tm["grupo"])
            ws.cell(row=row_n, column=2, value=tm["sigla"])
            ws.cell(row=row_n, column=3, value=tm["pais"])
            ws.cell(row=row_n, column=4, value=j)
            ws.cell(row=row_n, column=5, value=code)
            ws.cell(row=row_n, column=6, value=desc)
            ws.cell(row=row_n, column=7, value=foil_type(desc))
            ws.cell(row=row_n, column=9, value=0)
            ws.row_dimensions[row_n].height = 20
            row_n += 1
    last_sel = row_n - 1
    status_validation(ws, "H", 2, last_sel)
    status_conditional(ws, "H", 2, last_sel)
    ws.column_dimensions["H"].width = 14
    ws.freeze_panes = "E2"
    attach_table(ws, f"A1:I{last_sel}", "tblSelecoes", stripe=True)
    band_grupos(ws, last_sel)

    wm = wb.create_sheet("Museu FIFA")
    tab(wm, "7C3AED")
    heads = [H["idx"], H["code"], H["sticker"], H["typ"], H["status"], H["dup"]]
    for c, h in enumerate(heads, 1):
        wm.cell(row=1, column=c, value=h)
    apply_header(wm, 1, len(heads))
    for i, (code, desc) in enumerate(museum, start=1):
        wm.cell(row=i + 1, column=1, value=i)
        wm.cell(row=i + 1, column=2, value=code)
        wm.cell(row=i + 1, column=3, value=desc)
        wm.cell(row=i + 1, column=4, value=foil_type(desc))
        wm.cell(row=i + 1, column=6, value=0)
    last_mu = 1 + len(museum)
    status_validation(wm, "E", 2, last_mu)
    status_conditional(wm, "E", 2, last_mu)
    wm.freeze_panes = "E2"
    attach_table(wm, f"A1:F{last_mu}", "tblMuseu", stripe=True)

    wc = wb.create_sheet("Coca-Cola")
    tab(wc, "DC2626")
    heads = [H["promo"], H["player"], H["nat"], H["src"], H["status"], H["dup"]]
    for c, h in enumerate(heads, 1):
        wc.cell(row=1, column=c, value=h)
    apply_header(wc, 1, len(heads))
    for i, (num, player, sel) in enumerate(COCA_COLA, start=2):
        wc.cell(row=i, column=1, value=num)
        wc.cell(row=i, column=2, value=player)
        wc.cell(row=i, column=3, value=sel)
        wc.cell(row=i, column=4, value=ch("Rótulo EUA", "USA label"))
        wc.cell(row=i, column=6, value=0)
    last_cc = 1 + len(COCA_COLA)
    status_validation(wc, "E", 2, last_cc)
    status_conditional(wc, "E", 2, last_cc)
    wc.freeze_panes = "E2"
    attach_table(wc, f"A1:F{last_cc}", "tblCocaCola", stripe=True)

    we = wb.create_sheet("Extras")
    tab(we, "EA580C")
    heads = [H["idx"], H["player"], H["nat"], H["notes"], H["status"], H["dup"]]
    for c, h in enumerate(heads, 1):
        we.cell(row=1, column=c, value=h)
    apply_header(we, 1, len(heads))
    for i, (player, sel) in enumerate(EXTRAS, start=2):
        we.cell(row=i, column=1, value=i - 1)
        we.cell(row=i, column=2, value=player)
        we.cell(row=i, column=3, value=sel)
        we.cell(row=i, column=4, value=ch("Extra internacional", "International extra"))
        we.cell(row=i, column=6, value=0)
    last_ex = 1 + len(EXTRAS)
    status_validation(we, "E", 2, last_ex)
    status_conditional(we, "E", 2, last_ex)
    we.freeze_panes = "E2"
    attach_table(we, f"A1:F{last_ex}", "tblExtras", stripe=True)

    wx = wb.create_sheet("Siglas")
    tab(wx, "64748B")
    heads = [H["fifa"], H["country"], H["grp"], H["stk"], H["obs"]]
    for c, h in enumerate(heads, 1):
        wx.cell(row=1, column=c, value=h)
    apply_header(wx, 1, len(heads))
    for i, tm in enumerate(teams_meta, start=2):
        wx.cell(row=i, column=1, value=tm["sigla"])
        wx.cell(row=i, column=2, value=tm["pais"])
        wx.cell(row=i, column=3, value=tm["grupo"])
        wx.cell(row=i, column=4, value=20)
        if tm["sigla_panini"] != tm["sigla"]:
            wx.cell(row=i, column=5, value=f'Panini «{tm["sigla_panini"]}»')
    wx.freeze_panes = "A2"
    wx.auto_filter.ref = wx.dimensions
    autosize_columns(wx)

    wg = wb.create_sheet("Grupos")
    tab(wg, "475569")
    heads = [H["grp"], H["t1"], H["t2"], H["t3"], H["t4"]]
    for c, h in enumerate(heads, 1):
        wg.cell(row=1, column=c, value=h)
    apply_header(wg, 1, len(heads))
    for gi in range(12):
        slice_t = teams_meta[gi * 4 : (gi + 1) * 4]
        row = gi + 2
        wg.cell(row=row, column=1, value=f"Grupo {chr(ord('A') + gi)}")
        for j, tm in enumerate(slice_t):
            wg.cell(row=row, column=2 + j, value=f"{tm['pais']} ({tm['sigla']})")
    wg.freeze_panes = "A2"
    autosize_columns(wg)

    wa = wb.create_sheet("Todas (980)")
    tab(wa, "0369A1")
    heads = [H["sec"], H["grp"], H["abbr"], H["nat"], H["code"], H["sticker"], H["typ"], H["status"], H["dup"]]
    for c, h in enumerate(heads, 1):
        wa.cell(row=1, column=c, value=h)
    apply_header(wa, 1, len(heads))
    rr = 2
    for code, desc in intro:
        wa.cell(row=rr, column=1, value=ch("Página inicial", "Intro"))
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    for tm in teams_meta:
        for code, desc in tm["block"]:
            wa.cell(row=rr, column=1, value=ch("Seleção", "Team"))
            wa.cell(row=rr, column=2, value=tm["grupo"])
            wa.cell(row=rr, column=3, value=tm["sigla"])
            wa.cell(row=rr, column=4, value=tm["pais"])
            wa.cell(row=rr, column=5, value=code)
            wa.cell(row=rr, column=6, value=desc)
            wa.cell(row=rr, column=7, value=foil_type(desc))
            wa.cell(row=rr, column=9, value=0)
            rr += 1
    for code, desc in museum:
        wa.cell(row=rr, column=1, value=ch("Museu FIFA", "FIFA Museum"))
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    status_validation(wa, "H", 2, rr - 1)
    status_conditional(wa, "H", 2, rr - 1)
    wa.freeze_panes = "E2"
    attach_table(wa, f"A1:I{rr - 1}", "tblTodas980", stripe=True)
    autosize_columns(wa)


def build_by_nation(wb: Workbook, teams_meta) -> None:
    wp = wb.create_sheet("Por seleção")
    tab(wp, THEME["mint"])
    heads = [
        H["idx"],
        H["nat"],
        H["abbr"],
        H["grp"],
        H["goal"],
        H["got"],
        H["pct"],
        H["prog"],
        H["left"],
    ]
    for c, h in enumerate(heads, 1):
        wp.cell(row=1, column=c, value=h)
    apply_header(wp, 1, len(heads))
    for i, tm in enumerate(teams_meta, start=2):
        wp.cell(row=i, column=1, value=i - 1)
        wp.cell(row=i, column=2, value=tm["pais"])
        wp.cell(row=i, column=3, value=tm["sigla"])
        wp.cell(row=i, column=4, value=tm["grupo"])
        wp.cell(row=i, column=5, value=20)
        wp.cell(
            row=i,
            column=6,
            value=(
                f"=SUM(COUNTIFS(tblSelecoes[{H['abbr']}],C{i},tblSelecoes[{H['status']}],\"Sim\"),"
                f"COUNTIFS(tblSelecoes[{H['abbr']}],C{i},tblSelecoes[{H['status']}],\"Yes\"))"
            ),
        )
        wp.cell(row=i, column=7, value=f"=IF(E{i}>0,F{i}/E{i},0)")
        wp.cell(
            row=i,
            column=8,
            value=(
                f"=REPT(\"█\",MIN(14,MAX(0,ROUND(G{i}*14,0))))"
                f"&REPT(\"░\",14-MIN(14,MAX(0,ROUND(G{i}*14,0))))"
            ),
        )
        wp.cell(row=i, column=9, value=f"=E{i}-F{i}")
        wp.cell(row=i, column=7).number_format = "0%"
        wp.row_dimensions[i].height = 21
    attach_table(wp, f"A1:I{1 + len(teams_meta)}", "tblPorSelecao", stripe=True)
    wp.freeze_panes = "F2"
    autosize_columns(wp, max_width=42)


def build_packs(wb: Workbook) -> None:
    wp = wb.create_sheet("Pacotes")
    tab(wp, "A855F7")
    heads = [H["date"], H["packs"], H["new_est"], H["notes"]]
    for c, h in enumerate(heads, 1):
        wp.cell(row=1, column=c, value=h)
    apply_header(wp, 1, len(heads))
    for r in range(2, 42):
        wp.row_dimensions[r].height = 20
    attach_table(wp, "A1:D41", "tblPacotes", stripe=True)
    wp.freeze_panes = "A2"


def sumproduct_foil(tbl: str) -> str:
    st = H["status"]
    ty = H["typ"]
    return (
        f"SUMPRODUCT(({tbl}[{ty}]=\"FOIL\")"
        f"*(({tbl}[{st}]=\"Sim\")+({tbl}[{st}]=\"Yes\")))"
    )


def sumproduct_base(tbl: str) -> str:
    st = H["status"]
    ty = H["typ"]
    return (
        f"SUMPRODUCT(({tbl}[{ty}]=\"Base\")"
        f"*(({tbl}[{st}]=\"Sim\")+({tbl}[{st}]=\"Yes\")))"
    )


def build_stats(wb: Workbook) -> dict:
    ws = wb.create_sheet("Estatísticas")
    tab(ws, THEME["mint"])
    ws.merge_cells("A1:K1")
    ws["A1"] = ch("Panini WC 2026 · Motor", "Panini WC 2026 · Analytics")
    ws["A1"].font = Font(size=15, bold=True, color=THEME["header"])

    ws["A3"] = ch("Secção", "Section")
    ws["B3"] = ch("Meta", "Goal")
    ws["C3"] = ch("Com OK", "Got")
    ws["D3"] = "%"
    ws["E3"] = ch("Faltam", "Missing")
    for c in range(1, 6):
        ws.cell(row=3, column=c).font = Font(bold=True)
        ws.cell(row=3, column=c).fill = PatternFill("solid", fgColor="E8EEF7")

    rows = [
        (ch("Página inicial", "Intro"), 9, cnt_yes("tblPaginaInicial")),
        (ch("Seleções", "Teams"), 960, cnt_yes("tblSelecoes")),
        (ch("Museu FIFA", "Museum"), 11, cnt_yes("tblMuseu")),
    ]
    r = 4
    for lbl_cell, meta, cy in rows:
        ws.cell(row=r, column=1, value=lbl_cell)
        ws.cell(row=r, column=2, value=meta)
        ws.cell(row=r, column=3, value=f"={cy}")
        ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
        ws.cell(row=r, column=5, value=f"=B{r}-C{r}")
        ws.cell(row=r, column=4).number_format = "0.0%"
        r += 1

    ws.cell(row=r, column=1, value=ch("ÁLBUM BASE", "BASE ALBUM"))
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=2, value=980)
    ws.cell(
        row=r,
        column=3,
        value=f"={cnt_yes('tblPaginaInicial')}+{cnt_yes('tblSelecoes')}+{cnt_yes('tblMuseu')}",
    )
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=5, value=f"=B{r}-C{r}")
    ws.cell(row=r, column=4).number_format = "0.00%"
    row_album_total = r
    rsf, rsl = 4, 6

    r += 2
    ws.cell(row=r, column=1, value=ch("Promos & extras", "Promos & extras"))
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value=ch("Item", "Item"))
    ws.cell(row=r, column=2, value=ch("Meta", "Goal"))
    ws.cell(row=r, column=3, value=ch("Com OK", "Got"))
    ws.cell(row=r, column=4, value="%")
    for c in range(1, 5):
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Coca-Cola")
    ws.cell(row=r, column=2, value=len(COCA_COLA))
    ws.cell(row=r, column=3, value=f"={cnt_yes('tblCocaCola')}")
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=4).number_format = "0%"
    r += 1
    ws.cell(row=r, column=1, value=ch("Extras intl.", "Intl. extras"))
    ws.cell(row=r, column=2, value=len(EXTRAS))
    ws.cell(row=r, column=3, value=f"={cnt_yes('tblExtras')}")
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=4).number_format = "0%"
    row_ext_end = r

    r = row_ext_end + 3
    ws.cell(row=r, column=1, value=ch("FOIL vs Base", "FOIL vs Base"))
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="FOIL")
    ws.cell(
        row=r,
        column=2,
        value=f"={sumproduct_foil('tblPaginaInicial')}+{sumproduct_foil('tblSelecoes')}+{sumproduct_foil('tblMuseu')}",
    )
    r += 1
    ws.cell(row=r, column=1, value="Base")
    ws.cell(
        row=r,
        column=2,
        value=f"={sumproduct_base('tblPaginaInicial')}+{sumproduct_base('tblSelecoes')}+{sumproduct_base('tblMuseu')}",
    )
    row_foil_end = r

    r = row_foil_end + 3
    ws.cell(row=r, column=1, value=ch("Por grupo", "By group"))
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    gh = r
    # Cabeçalhos Excel Table têm de ser texto estável (não fórmulas CHOOSE).
    ws.cell(row=r, column=1, value=H["grp"])
    ws.cell(row=r, column=2, value=H["goal"])
    ws.cell(row=r, column=3, value=H["got"])
    ws.cell(row=r, column=4, value=H["pct"])
    ws.cell(row=r, column=5, value=H["miss"])
    ws.cell(row=r, column=6, value=H["bar"])
    ws.cell(row=r, column=7, value="Rank")
    for c in range(1, 8):
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="D6E4FF")
    r += 1
    gf = r
    for gi in range(12):
        gname = f"Grupo {chr(ord('A') + gi)}"
        row = r + gi
        ws.cell(row=row, column=1, value=gname)
        ws.cell(row=row, column=2, value=80)
        ws.cell(
            row=row,
            column=3,
            value=(
                f"=SUM(COUNTIFS(tblSelecoes[{H['status']}],\"Sim\",tblSelecoes[{H['grp']}],\"{gname}\"),"
                f"COUNTIFS(tblSelecoes[{H['status']}],\"Yes\",tblSelecoes[{H['grp']}],\"{gname}\"))"
            ),
        )
        ws.cell(row=row, column=4, value=f"=IF(B{row}>0,C{row}/B{row},0)")
        ws.cell(row=row, column=5, value=f"=B{row}-C{row}")
        ws.cell(row=row, column=4).number_format = "0%"
        ws.cell(
            row=row,
            column=6,
            value=f'=REPT("█",ROUND(D{row}*12,0))&REPT("·",12-ROUND(D{row}*12,0))',
        )
        ws.cell(
            row=row,
            column=7,
            value=f"=RANK.EQ(D{row},D${gf}:D${gf + 11},0)",
        )
    gl = r + 11
    attach_table(ws, f"A{gh}:G{gl}", "tblPorGrupo", stripe=True)

    r = gl + 3
    ws.cell(row=r, column=1, value=ch("Soma dup.", "Dup sum"))
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(
        row=r,
        column=2,
        value=(
            f"=SUM(tblPaginaInicial[{H['dup']}])+SUM(tblSelecoes[{H['dup']}])"
            f"+SUM(tblMuseu[{H['dup']}])+SUM(tblCocaCola[{H['dup']}])+SUM(tblExtras[{H['dup']}])"
        ),
    )

    autosize_columns(ws)
    return {"rt": row_album_total, "gf": gf, "gl": gl, "rsf": rsf, "rsl": rsl}


def build_dashboard(wb: Workbook, stats: dict) -> None:
    wd = wb["Dashboard"]
    tab(wd, THEME["gold"])
    rt, gf, gl, rsf, rsl = stats["rt"], stats["gf"], stats["gl"], stats["rsf"], stats["rsl"]
    sh = "Estatísticas"

    for row in range(1, 48):
        for col in range(1, 24):
            wd.cell(row=row, column=col).fill = PatternFill("solid", fgColor=THEME["night"])
    paint_dash_mat(wd)
    wd.sheet_view.showGridLines = False

    wd.merge_cells("A1:W2")
    wd["A1"] = ch("Painel da coleção", "Collection dashboard")
    wd["A1"].font = Font(size=26, bold=True, color=THEME["cream"])

    wd.merge_cells("A3:W3")
    wd["A3"] = ch(
        "Dados das folhas de cromos (coluna Status). Idioma dos textos: Config · B2 (lista em G1:G2).",
        "Numbers come from sticker sheets («Status» column). UI language: Config · B2 (see G1:G2).",
    )
    wd["A3"].font = Font(size=10, color=THEME["muted"])

    sc = sq(sh, f"C{rt}")
    sb = sq(sh, f"B{rt}")
    se = sq(sh, f"E{rt}")
    sd = sq(sh, f"D{rt}")

    cards = [
        ("B5:D5", "B6:D6", ch("Álbum base", "Base album"), f"={sc}&\" / \"&{sb}"),
        ("E5:G5", "E6:G6", ch("Faltam", "Missing"), f"={se}"),
        ("H5:J5", "H6:J6", ch("% Completo", "% Done"), f"={sd}"),
        (
            "K5:M5",
            "K6:M6",
            "Coca-Cola",
            f"={cnt_yes('tblCocaCola')}&\" / \"&{len(COCA_COLA)}",
        ),
        (
            "N5:P5",
            "N6:P6",
            "Extras",
            f"={cnt_yes('tblExtras')}&\" / \"&{len(EXTRAS)}",
        ),
        (
            "Q5:S5",
            "Q6:S6",
            ch("Duplicatas", "Duplicates"),
            f"=SUM(tblPaginaInicial[{H['dup']}])+SUM(tblSelecoes[{H['dup']}])+SUM(tblMuseu[{H['dup']}])"
            f"+SUM(tblCocaCola[{H['dup']}])+SUM(tblExtras[{H['dup']}])",
        ),
        (
            "B8:D8",
            "B9:D9",
            ch("Precisa troca", "Need swap"),
            f"={cnt_trade('tblPaginaInicial')}+{cnt_trade('tblSelecoes')}+{cnt_trade('tblMuseu')}",
        ),
        (
            "E8:G8",
            "E9:G9",
            ch("FOIL OK", "FOIL OK"),
            f"={sumproduct_foil('tblPaginaInicial')}+{sumproduct_foil('tblSelecoes')}+{sumproduct_foil('tblMuseu')}",
        ),
        (
            "H8:J8",
            "H9:J9",
            ch("Equipas fechadas", "Nations done"),
            f"=COUNTIF(tblPorSelecao[{H['left']}],0)",
        ),
        ("K8:M8", "K9:M9", ch("Hoje", "Today"), "=TODAY()"),
    ]

    for rtit, rval, title_f, form in cards:
        wd.merge_cells(rtit)
        wd.merge_cells(rval)
        tl, vl = rtit.split(":")[0], rval.split(":")[0]
        wd[tl].value = title_f
        wd[tl].font = Font(size=10, color=THEME["muted"])
        wd[tl].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wd[vl].value = form
        wd[vl].alignment = Alignment(horizontal="center", vertical="center")
        kpi_gold_left(wd, tl)
        wd[vl].font = Font(size=16, bold=True, color=THEME["cream"])

    wd["H6"].number_format = "0.0%"
    wd["H6"].font = Font(size=26, bold=True, color=THEME["mint"])
    wd["K9"].number_format = "dddd dd/mm/yyyy"
    wd["K9"].font = Font(size=13, bold=True, color=THEME["cream"])

    wd["U40"] = ch("Coladas", "Collected")
    wd["V40"] = f"={sc}"
    wd["U41"] = ch("Faltam", "Missing")
    wd["V41"] = f"={se}"

    dough = DoughnutChart()
    dough.title = "WC 2026"
    dough.style = 26
    dough.add_data(Reference(wd, min_col=22, min_row=40, max_row=41), titles_from_data=False)
    dough.set_categories(Reference(wd, min_col=21, min_row=40, max_row=41))
    dough.dataLabels = DataLabelList()
    dough.dataLabels.showPercent = True
    wd.add_chart(dough, "B12")

    colc = BarChart()
    colc.type = "col"
    colc.style = 11
    colc.title = "Sections"
    colc.y_axis.title = "n"
    colc.add_data(Reference(wb[sh], min_col=3, min_row=rsf, max_row=rsl), titles_from_data=False)
    colc.set_categories(Reference(wb[sh], min_col=1, min_row=rsf, max_row=rsl))
    wd.add_chart(colc, "B28")

    bars = BarChart()
    bars.type = "bar"
    bars.style = 12
    bars.title = "% group"
    bars.add_data(Reference(wb[sh], min_col=4, min_row=gf, max_row=gl), titles_from_data=False)
    bars.set_categories(Reference(wb[sh], min_col=1, min_row=gf, max_row=gl))
    wd.add_chart(bars, "M12")

    wd.merge_cells("B42:W44")
    wd["B42"] = ch(
        "Por seleção: barra por país. Pacotes não entram nos totais.",
        "By nation: progress bars. Packs sheet does not affect totals.",
    )
    wd["B42"].font = Font(size=10, color=THEME["muted"])



def main() -> None:
    intro, museum, teams_meta = load_album(CHECKLIST_DEFAULT)
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Config", 0)
    wb.create_sheet("Guia", 1)
    wb.create_sheet("Dashboard", 2)

    build_config(wb)
    build_guide(wb)
    build_data(wb, intro, museum, teams_meta)
    build_by_nation(wb, teams_meta)
    build_packs(wb)
    stats = build_stats(wb)
    build_dashboard(wb, stats)

    wb.save(OUT_V5)
    print(f"Escrito: {OUT_V5}")


if __name__ == "__main__":
    main()
