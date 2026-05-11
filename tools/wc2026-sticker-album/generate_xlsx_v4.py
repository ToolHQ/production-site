#!/usr/bin/env python3
"""
Versão 4 — idioma (pt-BR / en-US), layouts refinados, fórmulas e folhas dinâmicas.

  python3 generate_xlsx_v4.py --lang pt-BR
  python3 generate_xlsx_v4.py --lang en-US
  WC2026_LANG=en-US python3 generate_xlsx_v4.py
"""
from __future__ import annotations

import argparse
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from generate_xlsx_v2 import attach_table, autosize_columns
from wc2026_album_core import CHECKLIST_DEFAULT, COCA_COLA, EXTRAS, load_album
from wc2026_locale import LocalePack, esc_formula, get_locale

DIR = Path(__file__).resolve().parent

THEME = {
    "night": "0F172A",
    "panel": "1E293B",
    "mat": "151F32",
    "header": "0F2744",
    "gold": "CA8A04",
    "gold_bar": "D4A012",
    "mint": "34D399",
    "cream": "FEFCE8",
    "muted": "94A3B8",
    "band": "F1F5F9",
}
COL_SIM = "C6EFCE"
COL_NAO = "F8D7DA"
COL_TROCA = "FFF3CD"
THIN = Side(style="thin", color="334155")


def T(tbl: str, col_header: str) -> str:
    """Referência estruturada Excel (col_header pode ter espaços)."""
    return f"{tbl}[{col_header}]"


def sq(sheet_name: str, cell: str) -> str:
    safe = sheet_name.replace("'", "''")
    return f"'{safe}'!{cell}"


def foil_type(desc: str) -> str:
    d = desc.strip()
    return "FOIL" if d.endswith("FOIL") else "Base"


def apply_header_v4(ws, row: int, cols: int) -> None:
    gold_b = Border(
        left=THIN,
        right=THIN,
        top=THIN,
        bottom=Side(style="thick", color=THEME["gold"]),
    )
    fill = PatternFill("solid", fgColor=THEME["header"])
    font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = gold_b
    ws.row_dimensions[row].height = 28


def tab(ws, color_hex: str) -> None:
    ws.sheet_properties.tabColor = color_hex.replace("#", "")


def add_status_validation(ws, col_letter: str, sr: int, er: int, L: LocalePack) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    opts = f'"{L.st_yes},{L.st_no},{L.st_trade}"'
    dv = DataValidation(type="list", formula1=opts, allow_blank=True)
    dv.errorTitle = "Invalid" if L.code.startswith("en") else "Entrada"
    dv.error = "Pick a value from the list." if L.code.startswith("en") else "Use a lista."
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{sr}:{col_letter}{er}")


def add_status_conditional(ws, col_letter: str, sr: int, er: int, L: LocalePack) -> None:
    r = f"{col_letter}{sr}:{col_letter}{er}"
    col = f"${col_letter}{sr}"
    yes, no, trade = esc_formula(L.st_yes), esc_formula(L.st_no), esc_formula(L.st_trade)
    ws.conditional_formatting.add(
        r,
        FormulaRule(formula=[f'{col}="{yes}"'], fill=PatternFill("solid", fgColor=COL_SIM.replace("#", ""))),
    )
    ws.conditional_formatting.add(
        r,
        FormulaRule(formula=[f'{col}="{no}"'], fill=PatternFill("solid", fgColor=COL_NAO.replace("#", ""))),
    )
    ws.conditional_formatting.add(
        r,
        FormulaRule(formula=[f'{col}="{trade}"'], fill=PatternFill("solid", fgColor=COL_TROCA.replace("#", ""))),
    )


def band_grupos(ws, last_row: int) -> None:
    rng = f"A2:{get_column_letter(ws.max_column)}{last_row}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=["=MOD(INT((ROW()-2)/80),2)=0"],
            fill=PatternFill("solid", fgColor=THEME["band"]),
        ),
    )


def hl(sheet_name: str, cell_a1: str, label: str) -> str:
    safe = sheet_name.replace("'", "''")
    lab = esc_formula(label)
    return f"=HYPERLINK(\"#'{safe}'!{cell_a1}\",\"{lab}\")"


def paint_dashboard_mat(wd, row_end: int = 46, col_end: int = 23) -> None:
    """Área central mais clara (tapete) para contraste com cartões."""
    for row in range(4, min(row_end, 40)):
        for col in range(2, col_end):
            wd.cell(row=row, column=col).fill = PatternFill("solid", fgColor=THEME["mat"])


def kpi_border_left(wd, top_left: str) -> None:
    c = wd[top_left]
    c.border = Border(
        left=Side(style="thick", color=THEME["gold"]),
        top=THIN,
        right=THIN,
        bottom=THIN,
    )


def build_guide(wb: Workbook, L: LocalePack) -> None:
    wg = wb[L.sh_guide]
    wg.sheet_view.showGridLines = False
    wg.merge_cells("A1:K1")
    wg.row_dimensions[1].height = 11
    wg["A1"].fill = PatternFill("solid", fgColor=THEME["gold_bar"])

    for r in range(2, 40):
        for c in range(1, 12):
            wg.cell(row=r, column=c).fill = PatternFill("solid", fgColor=THEME["night"])

    wg.merge_cells("A3:K7")
    wg["A3"] = L.guide_title
    wg["A3"].font = Font(name="Calibri", size=30, bold=True, color=THEME["cream"])
    wg["A3"].alignment = Alignment(horizontal="left", vertical="center")

    wg.merge_cells("A8:K11")
    wg["A8"] = L.guide_sub
    wg["A8"].font = Font(size=11, color=THEME["muted"])
    wg["A8"].alignment = Alignment(wrap_text=True, vertical="center")

    wg["A13"] = L.guide_map
    wg["A13"].font = Font(size=15, bold=True, color=THEME["cream"])

    nav = [
        (L.dashboard_title, L.sh_dashboard, "► " + L.dashboard_title),
        (L.sh_intro, L.sh_intro, "► " + L.sh_intro),
        (L.sh_teams, L.sh_teams, "► " + L.sh_teams),
        (L.sh_museum, L.sh_museum, "► " + L.sh_museum),
        (L.sh_coca, L.sh_coca, "► " + L.sh_coca),
        (L.sh_extras, L.sh_extras, "► " + L.sh_extras),
        (L.sh_by_nation, L.sh_by_nation, "► " + L.sh_by_nation),
        (L.sh_packs, L.sh_packs, "► " + L.sh_packs),
        (L.sh_stats, L.sh_stats, "► " + L.sh_stats),
        (L.sh_codes, L.sh_codes, "► " + L.sh_codes),
        (L.sh_groups, L.sh_groups, "► " + L.sh_groups),
        (L.sh_master, L.sh_master, "► " + L.sh_master),
    ]

    r = 15
    wg.cell(row=r, column=1, value=L.guide_section_col)
    wg.cell(row=r, column=2, value=L.guide_open_col)
    for c in (1, 2):
        wg.cell(row=r, column=c).font = Font(bold=True, color=THEME["cream"])
        wg.cell(row=r, column=c).fill = PatternFill("solid", fgColor=THEME["panel"])
    r += 1
    for lbl, sh, link_title in nav:
        wg.cell(row=r, column=1, value=lbl)
        wg.cell(row=r, column=2, value=hl(sh, "A1", link_title))
        wg.cell(row=r, column=1).font = Font(color=THEME["muted"])
        wg.cell(row=r, column=2).font = Font(color=THEME["mint"], underline="single")
        r += 1

    r += 1
    wg.cell(row=r, column=1, value=L.guide_legend)
    wg.cell(row=r, column=1).font = Font(bold=True, color=THEME["cream"])
    r += 1
    legend_desc = {
        L.st_yes: L.st_yes + " — OK" if not L.code.startswith("en") else "In the album",
        L.st_no: L.card_missing,
        L.st_trade: L.card_trade,
    }
    for lab in (L.st_yes, L.st_no, L.st_trade):
        wg.cell(row=r, column=1, value=lab)
        wg.cell(row=r, column=2, value=legend_desc[lab])
        fill_k = COL_SIM if lab == L.st_yes else COL_NAO if lab == L.st_no else COL_TROCA
        wg.cell(row=r, column=1).fill = PatternFill("solid", fgColor=fill_k.replace("#", ""))
        r += 1

    wg.column_dimensions["A"].width = 32
    wg.column_dimensions["B"].width = 56
    tab(wg, THEME["gold"])


def build_about(wb: Workbook, L: LocalePack) -> None:
    """Notas estáticas — não interfere com tabelas."""
    name = "Sobre" if L.code.startswith("pt") else "About"
    ws = wb.create_sheet(name)
    tab(ws, "64748B")
    ws["A1"] = "Panini FIFA World Cup 2026 — sticker workbook"
    ws["A1"].font = Font(bold=True, size=14, color=THEME["header"])
    ws["A3"] = f"Locale / Idioma: {L.label_human} ({L.code})"
    ws["A4"] = (
        "Gerado por generate_xlsx_v4.py — listas e fórmulas seguem os cabeçalhos deste idioma."
        if L.code.startswith("pt")
        else "Built by generate_xlsx_v4.py — formulas follow this locale's table headers."
    )
    ws["A6"] = (
        "Não combine AutoFilter da folha com Table no mesmo intervalo (limitação Excel)."
        if L.code.startswith("pt")
        else "Do not mix worksheet AutoFilter with Excel Tables on the same range."
    )
    ws.column_dimensions["A"].width = 88


def build_data(wb: Workbook, L: LocalePack, intro, museum, teams_meta) -> None:
    w1 = wb.create_sheet(L.sh_intro)
    tab(w1, "2563EB")
    h1 = [L.h_num, L.h_code, L.h_desc, L.h_type, L.h_have, L.h_dup]
    for c, h in enumerate(h1, 1):
        w1.cell(row=1, column=c, value=h)
    apply_header_v4(w1, 1, len(h1))
    for i, (code, desc) in enumerate(intro, start=1):
        row = i + 1
        w1.cell(row=row, column=1, value=i)
        w1.cell(row=row, column=2, value=code)
        w1.cell(row=row, column=3, value=desc)
        w1.cell(row=row, column=4, value=foil_type(desc))
        w1.cell(row=row, column=6, value=0)
        w1.row_dimensions[row].height = 22
    last_pi = 1 + len(intro)
    add_status_validation(w1, "E", 2, last_pi, L)
    add_status_conditional(w1, "E", 2, last_pi, L)
    w1.column_dimensions["E"].width = 16
    w1.column_dimensions["C"].width = 44
    w1.freeze_panes = "E2"
    attach_table(w1, f"A1:F{last_pi}", "tblPaginaInicial", stripe=True)

    ws = wb.create_sheet(L.sh_teams)
    tab(ws, "059669")
    hs = [
        L.h_grp,
        L.h_abbr,
        L.h_team,
        L.h_num,
        L.h_code,
        L.h_desc,
        L.h_type,
        L.h_have,
        L.h_dup,
    ]
    for c, h in enumerate(hs, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header_v4(ws, 1, len(hs))
    row_n = 2
    for tm in teams_meta:
        for j, (code, desc) in enumerate(tm["block"], start=1):
            ws.cell(row=row_n, column=1, value=tm["grupo"])
            ws.cell(row=row_n, column=2, value=tm["sigla"])
            ws.cell(row=row_n, column=3, value=tm["pais"])
            ws.cell(row=row_n, column=4, value=j)
            ws.cell(row=row_n, column=5, value=code)
            ws.cell(row=row_n, column=6, value=desc)
            ws.cell(row=row_n, column=7, value=foil_type(desc))
            ws.cell(row=row_n, column=9, value=0)
            ws.row_dimensions[row_n].height = 20
            row_n += 1
    last_sel = row_n - 1
    add_status_validation(ws, "H", 2, last_sel, L)
    add_status_conditional(ws, "H", 2, last_sel, L)
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["F"].width = 38
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "E2"
    attach_table(ws, f"A1:I{last_sel}", "tblSelecoes", stripe=True)
    band_grupos(ws, last_sel)

    wm = wb.create_sheet(L.sh_museum)
    tab(wm, "7C3AED")
    hm = [L.h_num, L.h_code, L.h_desc, L.h_type, L.h_have, L.h_dup]
    for c, h in enumerate(hm, 1):
        wm.cell(row=1, column=c, value=h)
    apply_header_v4(wm, 1, len(hm))
    for i, (code, desc) in enumerate(museum, start=1):
        wm.cell(row=i + 1, column=1, value=i)
        wm.cell(row=i + 1, column=2, value=code)
        wm.cell(row=i + 1, column=3, value=desc)
        wm.cell(row=i + 1, column=4, value=foil_type(desc))
        wm.cell(row=i + 1, column=6, value=0)
        wm.row_dimensions[i + 1].height = 22
    last_mu = 1 + len(museum)
    add_status_validation(wm, "E", 2, last_mu, L)
    add_status_conditional(wm, "E", 2, last_mu, L)
    wm.column_dimensions["E"].width = 16
    wm.column_dimensions["C"].width = 42
    wm.freeze_panes = "E2"
    attach_table(wm, f"A1:F{last_mu}", "tblMuseu", stripe=True)

    wc = wb.create_sheet(L.sh_coca)
    tab(wc, "DC2626")
    ch = [L.h_promo, L.h_player, L.h_team, L.h_origin, L.h_have, L.h_dup]
    for c, h in enumerate(ch, 1):
        wc.cell(row=1, column=c, value=h)
    apply_header_v4(wc, 1, len(ch))
    for i, (num, player, sel) in enumerate(COCA_COLA, start=2):
        wc.cell(row=i, column=1, value=num)
        wc.cell(row=i, column=2, value=player)
        wc.cell(row=i, column=3, value=sel)
        wc.cell(row=i, column=4, value=L.txt_coca_origin)
        wc.cell(row=i, column=6, value=0)
        wc.row_dimensions[i].height = 22
    last_cc = 1 + len(COCA_COLA)
    add_status_validation(wc, "E", 2, last_cc, L)
    add_status_conditional(wc, "E", 2, last_cc, L)
    wc.column_dimensions["E"].width = 16
    wc.freeze_panes = "E2"
    attach_table(wc, f"A1:F{last_cc}", "tblCocaCola", stripe=True)

    we = wb.create_sheet(L.sh_extras)
    tab(we, "EA580C")
    eh = [L.h_num, L.h_player, L.h_team, L.h_notes, L.h_have, L.h_dup]
    for c, h in enumerate(eh, 1):
        we.cell(row=1, column=c, value=h)
    apply_header_v4(we, 1, len(eh))
    for i, (player, sel) in enumerate(EXTRAS, start=2):
        we.cell(row=i, column=1, value=i - 1)
        we.cell(row=i, column=2, value=player)
        we.cell(row=i, column=3, value=sel)
        we.cell(row=i, column=4, value=L.txt_extra_note)
        we.cell(row=i, column=6, value=0)
        we.row_dimensions[i].height = 22
    last_ex = 1 + len(EXTRAS)
    add_status_validation(we, "E", 2, last_ex, L)
    add_status_conditional(we, "E", 2, last_ex, L)
    we.column_dimensions["E"].width = 16
    we.freeze_panes = "E2"
    attach_table(we, f"A1:F{last_ex}", "tblExtras", stripe=True)

    wx = wb.create_sheet(L.sh_codes)
    tab(wx, "64748B")
    xh = [L.h_fifa_code, L.h_country_name, L.h_grp, L.h_stickers, L.h_obs]
    for c, h in enumerate(xh, 1):
        wx.cell(row=1, column=c, value=h)
    apply_header_v4(wx, 1, len(xh))
    for i, tm in enumerate(teams_meta, start=2):
        wx.cell(row=i, column=1, value=tm["sigla"])
        wx.cell(row=i, column=2, value=tm["pais"])
        wx.cell(row=i, column=3, value=tm["grupo"])
        wx.cell(row=i, column=4, value=20)
        if tm["sigla_panini"] != tm["sigla"]:
            wx.cell(row=i, column=5, value=f"{L.txt_panini_prefix} «{tm['sigla_panini']}»")
    wx.freeze_panes = "A2"
    wx.auto_filter.ref = wx.dimensions
    autosize_columns(wx)

    wg = wb.create_sheet(L.sh_groups)
    tab(wg, "475569")
    gh = [L.h_grp, L.h_pick1, L.h_pick2, L.h_pick3, L.h_pick4]
    for c, h in enumerate(gh, 1):
        wg.cell(row=1, column=c, value=h)
    apply_header_v4(wg, 1, len(gh))
    for gi in range(12):
        slice_t = teams_meta[gi * 4 : (gi + 1) * 4]
        row = gi + 2
        wg.cell(row=row, column=1, value=f"Grupo {chr(ord('A') + gi)}")
        for j, tm in enumerate(slice_t):
            wg.cell(row=row, column=2 + j, value=f"{tm['pais']} ({tm['sigla']})")
    wg.freeze_panes = "A2"
    autosize_columns(wg)

    wa = wb.create_sheet(L.sh_master)
    tab(wa, "0369A1")
    ah = [
        L.h_section,
        L.h_grp,
        L.h_abbr,
        L.h_team,
        L.h_code,
        L.h_desc,
        L.h_type,
        L.h_have,
        L.h_dup_short,
    ]
    for c, h in enumerate(ah, 1):
        wa.cell(row=1, column=c, value=h)
    apply_header_v4(wa, 1, len(ah))
    rr = 2
    for code, desc in intro:
        wa.cell(row=rr, column=1, value=L.txt_master_intro)
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    for tm in teams_meta:
        for code, desc in tm["block"]:
            wa.cell(row=rr, column=1, value=L.txt_master_team)
            wa.cell(row=rr, column=2, value=tm["grupo"])
            wa.cell(row=rr, column=3, value=tm["sigla"])
            wa.cell(row=rr, column=4, value=tm["pais"])
            wa.cell(row=rr, column=5, value=code)
            wa.cell(row=rr, column=6, value=desc)
            wa.cell(row=rr, column=7, value=foil_type(desc))
            wa.cell(row=rr, column=9, value=0)
            rr += 1
    for code, desc in museum:
        wa.cell(row=rr, column=1, value=L.txt_master_museum)
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    add_status_validation(wa, "H", 2, rr - 1, L)
    add_status_conditional(wa, "H", 2, rr - 1, L)
    wa.freeze_panes = "E2"
    wa.column_dimensions["H"].width = 16
    attach_table(wa, f"A1:I{rr - 1}", "tblTodas980", stripe=True)
    autosize_columns(wa)


def build_by_nation(wb: Workbook, L: LocalePack, teams_meta) -> None:
    wp = wb.create_sheet(L.sh_by_nation)
    tab(wp, THEME["mint"])
    heads = [
        L.h_num,
        L.h_team,
        L.h_abbr,
        L.h_grp,
        L.h_goal,
        L.h_stuck,
        L.h_pct,
        L.h_prog,
        L.h_left,
    ]
    for c, h in enumerate(heads, 1):
        wp.cell(row=1, column=c, value=h)
    apply_header_v4(wp, 1, len(heads))
    sy = esc_formula(L.st_yes)
    for i, tm in enumerate(teams_meta, start=2):
        wp.cell(row=i, column=1, value=i - 1)
        wp.cell(row=i, column=2, value=tm["pais"])
        wp.cell(row=i, column=3, value=tm["sigla"])
        wp.cell(row=i, column=4, value=tm["grupo"])
        wp.cell(row=i, column=5, value=20)
        wp.cell(
            row=i,
            column=6,
            value=(
                f'=COUNTIFS(tblSelecoes[{L.h_abbr}],C{i},'
                f'tblSelecoes[{L.h_have}],"{sy}")'
            ),
        )
        wp.cell(row=i, column=7, value=f"=IF(E{i}>0,F{i}/E{i},0)")
        wp.cell(
            row=i,
            column=8,
            value=(
                f'=REPT("█",MIN(14,MAX(0,ROUND(G{i}*14,0))))'
                f'&REPT("░",14-MIN(14,MAX(0,ROUND(G{i}*14,0))))'
            ),
        )
        wp.cell(row=i, column=9, value=f"=E{i}-F{i}")
        wp.cell(row=i, column=7).number_format = "0%"
        wp.row_dimensions[i].height = 21
    attach_table(wp, f"A1:I{1 + len(teams_meta)}", "tblPorSelecao", stripe=True)
    wp.freeze_panes = "F2"
    wp.column_dimensions["B"].width = 26
    wp.column_dimensions["C"].width = 8
    wp.column_dimensions["D"].width = 11
    wp.column_dimensions["H"].width = 18
    autosize_columns(wp, max_width=40)


def build_packs(wb: Workbook, L: LocalePack) -> None:
    wp = wb.create_sheet(L.sh_packs)
    tab(wp, "A855F7")
    heads = [L.h_date, L.h_packs_opened, L.h_new_est, L.h_notes]
    for c, h in enumerate(heads, 1):
        wp.cell(row=1, column=c, value=h)
    apply_header_v4(wp, 1, len(heads))
    for r in range(2, 42):
        wp.row_dimensions[r].height = 20
    attach_table(wp, "A1:D41", "tblPacotes", stripe=True)
    wp.freeze_panes = "A2"
    wp.column_dimensions["A"].width = 14
    wp.column_dimensions["B"].width = 16
    wp.column_dimensions["C"].width = 22
    wp.column_dimensions["D"].width = 44


def build_stats(wb: Workbook, L: LocalePack) -> dict:
    ws = wb.create_sheet(L.sh_stats)
    tab(ws, THEME["mint"])
    sy = esc_formula(L.st_yes)
    tr = esc_formula(L.st_trade)

    ws.merge_cells("A1:K1")
    ws["A1"] = L.stats_title
    ws["A1"].font = Font(size=15, bold=True, color=THEME["header"])

    ws["A3"] = L.sec_item
    ws["B3"] = L.sec_meta
    ws["C3"] = L.sec_with_yes
    ws["D3"] = L.sec_pct
    ws["E3"] = L.sec_left
    for c in range(1, 6):
        ws.cell(row=3, column=c).font = Font(bold=True)
        ws.cell(row=3, column=c).fill = PatternFill("solid", fgColor="E8EEF7")

    cf_pi = f'COUNTIF(tblPaginaInicial[{L.h_have}],"{sy}")'
    cf_te = f'COUNTIF(tblSelecoes[{L.h_have}],"{sy}")'
    cf_mu = f'COUNTIF(tblMuseu[{L.h_have}],"{sy}")'

    rows_sec = [
        (L.sec_intro, 9, cf_pi),
        (L.sec_teams_block, 960, cf_te),
        (L.sec_museum, 11, cf_mu),
    ]
    r = 4
    for label, meta_val, fsim in rows_sec:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=meta_val)
        ws.cell(row=r, column=3, value=f"={fsim}")
        ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
        ws.cell(row=r, column=5, value=f"=B{r}-C{r}")
        ws.cell(row=r, column=4).number_format = "0.0%"
        r += 1

    ws.cell(row=r, column=1, value=L.sec_album_total)
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=2, value=980)
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(
        row=r,
        column=3,
        value=f"={cf_pi}+{cf_te}+{cf_mu}",
    )
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=5, value=f"=B{r}-C{r}")
    ws.cell(row=r, column=4).number_format = "0.00%"
    row_album_total = r
    row_sec_first = 4
    row_sec_last = 6

    r += 2
    ws.cell(row=r, column=1, value=L.promo_title)
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value=L.sec_item)
    ws.cell(row=r, column=2, value=L.sec_meta)
    ws.cell(row=r, column=3, value=L.sec_with_yes)
    ws.cell(row=r, column=4, value=L.sec_pct)
    for c in range(1, 5):
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value=L.sec_coca)
    ws.cell(row=r, column=2, value=len(COCA_COLA))
    ws.cell(row=r, column=3, value=f'=COUNTIF(tblCocaCola[{L.h_have}],"{sy}")')
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=4).number_format = "0%"
    r += 1
    ws.cell(row=r, column=1, value=L.sec_extras_lbl)
    ws.cell(row=r, column=2, value=len(EXTRAS))
    ws.cell(row=r, column=3, value=f'=COUNTIF(tblExtras[{L.h_have}],"{sy}")')
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=4).number_format = "0%"
    row_ext_end = r

    r = row_ext_end + 3
    ws.cell(row=r, column=1, value=L.foil_block)
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value=L.sec_foil)
    ws.cell(
        row=r,
        column=2,
        value=(
            f'=SUMPRODUCT((tblPaginaInicial[{L.h_type}]="FOIL")*(tblPaginaInicial[{L.h_have}]="{sy}"))'
            f'+SUMPRODUCT((tblSelecoes[{L.h_type}]="FOIL")*(tblSelecoes[{L.h_have}]="{sy}"))'
            f'+SUMPRODUCT((tblMuseu[{L.h_type}]="FOIL")*(tblMuseu[{L.h_have}]="{sy}"))'
        ),
    )
    r += 1
    ws.cell(row=r, column=1, value=L.sec_base)
    ws.cell(
        row=r,
        column=2,
        value=(
            f'=SUMPRODUCT((tblPaginaInicial[{L.h_type}]="Base")*(tblPaginaInicial[{L.h_have}]="{sy}"))'
            f'+SUMPRODUCT((tblSelecoes[{L.h_type}]="Base")*(tblSelecoes[{L.h_have}]="{sy}"))'
            f'+SUMPRODUCT((tblMuseu[{L.h_type}]="Base")*(tblMuseu[{L.h_have}]="{sy}"))'
        ),
    )
    row_foil_end = r

    r = row_foil_end + 3
    ws.cell(row=r, column=1, value=L.grp_progress)
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    grp_header_row = r
    ws.cell(row=r, column=1, value=L.h_grp)
    ws.cell(row=r, column=2, value=L.grp_meta)
    ws.cell(row=r, column=3, value=L.grp_have)
    ws.cell(row=r, column=4, value=L.grp_pct)
    ws.cell(row=r, column=5, value=L.grp_missing)
    ws.cell(row=r, column=6, value=L.grp_bar)
    ws.cell(row=r, column=7, value=L.grp_rank)
    for c in range(1, 8):
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="D6E4FF")
    r += 1
    grp_first_row = r

    for gi in range(12):
        gname = f"Grupo {chr(ord('A') + gi)}"
        row = r + gi
        ws.cell(row=row, column=1, value=gname)
        ws.cell(row=row, column=2, value=80)
        ws.cell(
            row=row,
            column=3,
            value=(
                f'=COUNTIFS(tblSelecoes[{L.h_have}],"{sy}",'
                f'tblSelecoes[{L.h_grp}],"{gname}")'
            ),
        )
        ws.cell(row=row, column=4, value=f"=IF(B{row}>0,C{row}/B{row},0)")
        ws.cell(row=row, column=5, value=f"=B{row}-C{row}")
        ws.cell(row=row, column=4).number_format = "0%"
        ws.cell(
            row=row,
            column=6,
            value=f'=REPT("█",ROUND(D{row}*12,0))&REPT("·",12-ROUND(D{row}*12,0))',
        )
        ws.cell(
            row=row,
            column=7,
            value=f"=RANK.EQ(D{row},D${grp_first_row}:D${grp_first_row + 11},0)",
        )
    grp_last_row = r + 11
    attach_table(ws, f"A{grp_header_row}:G{grp_last_row}", "tblPorGrupo", stripe=True)

    r = grp_last_row + 3
    ws.cell(row=r, column=1, value=L.dup_sum)
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(
        row=r,
        column=2,
        value=(
            f"=SUM(tblPaginaInicial[{L.h_dup}])+SUM(tblSelecoes[{L.h_dup}])"
            f"+SUM(tblMuseu[{L.h_dup}])+SUM(tblCocaCola[{L.h_dup}])+SUM(tblExtras[{L.h_dup}])"
        ),
    )

    autosize_columns(ws)

    return {
        "row_album_total": row_album_total,
        "grp_first_row": grp_first_row,
        "grp_last_row": grp_last_row,
        "row_sec_first": row_sec_first,
        "row_sec_last": row_sec_last,
        "trade_formula": tr,
        "sy": sy,
    }


def build_dashboard(wb: Workbook, L: LocalePack, stats: dict) -> None:
    wd = wb[L.sh_dashboard]
    tab(wd, THEME["gold"])
    rt = stats["row_album_total"]
    gf = stats["grp_first_row"]
    gl = stats["grp_last_row"]
    rsf = stats["row_sec_first"]
    rsl = stats["row_sec_last"]
    tr = stats["trade_formula"]
    sy = stats["sy"]
    sh_st = L.sh_stats

    for row in range(1, 48):
        for col in range(1, 24):
            wd.cell(row=row, column=col).fill = PatternFill("solid", fgColor=THEME["night"])
    paint_dashboard_mat(wd)
    wd.sheet_view.showGridLines = False

    wd.merge_cells("A1:W2")
    wd["A1"] = L.dashboard_title
    wd["A1"].font = Font(name="Calibri", size=26, bold=True, color=THEME["cream"])
    wd["A1"].alignment = Alignment(horizontal="left", vertical="center")

    wd.merge_cells("A3:W3")
    wd["A3"] = L.dashboard_sub
    wd["A3"].font = Font(size=10, color=THEME["muted"])
    wd["A3"].alignment = Alignment(wrap_text=True)

    stat_c_rt = sq(sh_st, f"C{rt}")
    stat_b_rt = sq(sh_st, f"B{rt}")
    stat_e_rt = sq(sh_st, f"E{rt}")
    stat_d_rt = sq(sh_st, f"D{rt}")

    cards = [
        ("B5:D5", "B6:D6", L.card_album, f"={stat_c_rt}&\" / \"&{stat_b_rt}"),
        ("E5:G5", "E6:G6", L.card_missing, f"={stat_e_rt}"),
        ("H5:J5", "H6:J6", L.card_pct, f"={stat_d_rt}"),
        ("K5:M5", "K6:M6", L.card_coca, f'=COUNTIF(tblCocaCola[{L.h_have}],"{sy}")&" / "&COUNTA(tblCocaCola[{L.h_have}])'),
        ("N5:P5", "N6:P6", L.card_extras, f'=COUNTIF(tblExtras[{L.h_have}],"{sy}")&" / "&COUNTA(tblExtras[{L.h_have}])'),
        (
            "Q5:S5",
            "Q6:S6",
            L.card_dup,
            f"=SUM(tblPaginaInicial[{L.h_dup}])+SUM(tblSelecoes[{L.h_dup}])+SUM(tblMuseu[{L.h_dup}])"
            f"+SUM(tblCocaCola[{L.h_dup}])+SUM(tblExtras[{L.h_dup}])",
        ),
        (
            "B8:D8",
            "B9:D9",
            L.card_trade,
            f'=COUNTIF(tblPaginaInicial[{L.h_have}],"{tr}")+COUNTIF(tblSelecoes[{L.h_have}],"{tr}")'
            f'+COUNTIF(tblMuseu[{L.h_have}],"{tr}")',
        ),
        (
            "E8:G8",
            "E9:G9",
            L.card_foil,
            f'=SUMPRODUCT((tblPaginaInicial[{L.h_type}]="FOIL")*(tblPaginaInicial[{L.h_have}]="{sy}"))'
            f'+SUMPRODUCT((tblSelecoes[{L.h_type}]="FOIL")*(tblSelecoes[{L.h_have}]="{sy}"))'
            f'+SUMPRODUCT((tblMuseu[{L.h_type}]="FOIL")*(tblMuseu[{L.h_have}]="{sy}"))',
        ),
        ("H8:J8", "H9:J9", L.card_done_teams, f"=COUNTIF(tblPorSelecao[{L.h_left}],0)"),
        ("K8:M8", "K9:M9", L.card_today, "=TODAY()"),
    ]

    for rng_title, rng_val, title, formula in cards:
        wd.merge_cells(rng_title)
        wd.merge_cells(rng_val)
        tl = rng_title.split(":")[0]
        vl = rng_val.split(":")[0]
        wd[tl].value = title
        wd[tl].font = Font(size=10, color=THEME["muted"])
        wd[tl].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cv = wd[vl]
        cv.value = formula
        cv.alignment = Alignment(horizontal="center", vertical="center")
        kpi_border_left(wd, tl)
        if title == L.card_pct:
            cv.font = Font(size=26, bold=True, color=THEME["mint"])
            cv.number_format = "0.0%"
        elif title == L.card_today:
            cv.font = Font(size=13, bold=True, color=THEME["cream"])
            cv.number_format = "dddd dd/mm/yyyy"
        else:
            cv.font = Font(size=16, bold=True, color=THEME["cream"])

    wd["U40"] = L.chart_have_slice
    wd["V40"] = f"={sq(sh_st, f'C{rt}')}"
    wd["U41"] = L.chart_miss_slice
    wd["V41"] = f"={sq(sh_st, f'E{rt}')}"

    dough = DoughnutChart()
    dough.title = L.chart_doughnut
    dough.style = 26
    dough.add_data(Reference(wd, min_col=22, min_row=40, max_row=41), titles_from_data=False)
    dough.set_categories(Reference(wd, min_col=21, min_row=40, max_row=41))
    dough.dataLabels = DataLabelList()
    dough.dataLabels.showPercent = True
    wd.add_chart(dough, "B12")

    colc = BarChart()
    colc.type = "col"
    colc.style = 11
    colc.title = L.chart_sections
    colc.y_axis.title = L.chart_stickers_axis
    colc.add_data(
        Reference(wb[sh_st], min_col=3, min_row=rsf, max_row=rsl),
        titles_from_data=False,
    )
    colc.set_categories(Reference(wb[sh_st], min_col=1, min_row=rsf, max_row=rsl))
    wd.add_chart(colc, "B28")

    bars = BarChart()
    bars.type = "bar"
    bars.style = 12
    bars.title = L.chart_groups_pct
    bars.x_axis.title = L.chart_pct_axis
    bars.add_data(Reference(wb[sh_st], min_col=4, min_row=gf, max_row=gl), titles_from_data=False)
    bars.set_categories(Reference(wb[sh_st], min_col=1, min_row=gf, max_row=gl))
    wd.add_chart(bars, "M12")

    wd.merge_cells("B42:W44")
    wd["B42"] = L.dashboard_tip
    wd["B42"].font = Font(size=10, color=THEME["muted"])
    wd["B42"].alignment = Alignment(wrap_text=True, vertical="top")

    for rr in (40, 41):
        wd.row_dimensions[rr].hidden = True

    wd.column_dimensions["B"].width = 12


def main() -> None:
    ap = argparse.ArgumentParser(description="Panini WC 2026 sticker workbook v4")
    ap.add_argument(
        "--lang",
        default=os.environ.get("WC2026_LANG", "pt-BR"),
        help="pt-BR | en-US (ou env WC2026_LANG)",
    )
    ap.add_argument("-o", "--output", type=Path, default=None, help="Caminho do .xlsx")
    args = ap.parse_args()

    L = get_locale(args.lang)
    intro, museum, teams_meta = load_album(CHECKLIST_DEFAULT)

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet(L.sh_guide, 0)
    wb.create_sheet(L.sh_dashboard, 1)

    build_guide(wb, L)
    build_data(wb, L, intro, museum, teams_meta)
    build_by_nation(wb, L, teams_meta)
    build_packs(wb, L)
    stats = build_stats(wb, L)
    build_dashboard(wb, L, stats)
    build_about(wb, L)

    out = args.output
    if out is None:
        suf = "pt" if L.code == "pt-BR" else "en"
        out = DIR / f"FIFA-World-Cup-2026-Panini-Controle-v4-{suf}.xlsx"
    wb.save(out)
    print(f"Escrito: {out} ({L.label_human})")


if __name__ == "__main__":
    main()
