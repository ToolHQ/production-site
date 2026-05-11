#!/usr/bin/env python3
"""
Versão 2: dashboard dinâmico, estatísticas, tabelas Excel e formatação condicional
para controle do álbum Panini FIFA World Cup 2026.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from wc2026_album_core import CHECKLIST_DEFAULT, COCA_COLA, EXTRAS, load_album

DIR = Path(__file__).resolve().parent
OUT_V2 = DIR / "FIFA-World-Cup-2026-Panini-Controle-v2.xlsx"

# --- Visual tokens (tema “nocaute”) ---
BG_DASH = "0B132B"
COL_SIM = "C6EFCE"
COL_NAO = "F8D7DA"
COL_TROCA = "FFF3CD"
THIN = Side(style="thin", color="1B263B")

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
DASH_TITLE = Font(bold=True, size=20, color="FFFFFF")
DASH_KPI = Font(bold=True, size=28, color="06D6A0")
DASH_LBL = Font(size=11, color="8B9BB4")
DASH_SUB = Font(size=9, color="5C6B8A")


def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()


def add_tenho_validation(ws, col_letter: str, start_row: int, end_row: int):
    dv = DataValidation(
        type="list",
        formula1='"Sim,Não,Falta trocar"',
        allow_blank=True,
    )
    dv.error = "Use a lista: Sim, Não ou Falta trocar."
    dv.errorTitle = "Entrada"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def add_tenho_conditional(ws, col_letter: str, start_row: int, end_row: int):
    """Realça a célula «Tenho» consoante o estado."""
    r = f"{col_letter}{start_row}:{col_letter}{end_row}"
    col = f"${col_letter}{start_row}"
    ws.conditional_formatting.add(
        r,
        FormulaRule(
            formula=[f'{col}="Sim"'],
            fill=PatternFill("solid", fgColor=COL_SIM.replace("#", "")),
        ),
    )
    ws.conditional_formatting.add(
        r,
        FormulaRule(
            formula=[f'{col}="Não"'],
            fill=PatternFill("solid", fgColor=COL_NAO.replace("#", "")),
        ),
    )
    ws.conditional_formatting.add(
        r,
        FormulaRule(
            formula=[f'{col}="Falta trocar"'],
            fill=PatternFill("solid", fgColor=COL_TROCA.replace("#", "")),
        ),
    )


def attach_table(ws, ref: str, name: str, stripe: bool = True):
    """Anexa tabela Excel com nome estável para referências estruturadas."""
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=stripe,
        showColumnStripes=False,
    )
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = style
    ws.add_table(tab)


def autosize_columns(ws, max_width: int = 48, min_width: float = 8):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        maxlen = 0
        for cell in col:
            if cell.value is None:
                continue
            maxlen = max(maxlen, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, maxlen + 2))


def foil_type(desc: str) -> str:
    d = desc.strip()
    return "FOIL" if d.endswith("FOIL") else "Base"


def build_data_sheets(wb: Workbook, intro, museum, teams_meta) -> dict:
    """Cria folhas de dados; devolve coordenadas para fórmulas do dashboard."""
    meta = {}

    # --- Página Inicial ---
    w1 = wb.create_sheet("Página Inicial")
    h1 = ["#", "Código", "Descrição no álbum", "Tipo", "Tenho", "Duplicatas"]
    for c, h in enumerate(h1, 1):
        w1.cell(row=1, column=c, value=h)
    apply_header(w1, 1, len(h1))
    for i, (code, desc) in enumerate(intro, start=1):
        row = i + 1
        w1.cell(row=row, column=1, value=i)
        w1.cell(row=row, column=2, value=code)
        w1.cell(row=row, column=3, value=desc)
        w1.cell(row=row, column=4, value=foil_type(desc))
        w1.cell(row=row, column=6, value=0)
        w1.row_dimensions[row].height = 22
    last_pi = 1 + len(intro)
    add_tenho_validation(w1, "E", 2, last_pi)
    add_tenho_conditional(w1, "E", 2, last_pi)
    w1.column_dimensions["E"].width = 16
    w1.column_dimensions["C"].width = 42
    w1.freeze_panes = "E2"
    attach_table(w1, f"A1:F{last_pi}", "tblPaginaInicial", stripe=True)

    # --- Seleções ---
    ws = wb.create_sheet("Seleções")
    hs = ["Grupo", "Sigla", "Seleção", "#", "Código", "Descrição no álbum", "Tipo", "Tenho", "Duplicatas"]
    for c, h in enumerate(hs, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header(ws, 1, len(hs))
    row_n = 2
    for tm in teams_meta:
        for j, (code, desc) in enumerate(tm["block"], start=1):
            ws.cell(row=row_n, column=1, value=tm["grupo"])
            ws.cell(row=row_n, column=2, value=tm["sigla"])
            ws.cell(row=row_n, column=3, value=tm["pais"])
            ws.cell(row=row_n, column=4, value=j)
            ws.cell(row=row_n, column=5, value=code)
            ws.cell(row=row_n, column=6, value=desc)
            ws.cell(row=row_n, column=7, value=foil_type(desc))
            ws.cell(row=row_n, column=9, value=0)
            ws.row_dimensions[row_n].height = 20
            row_n += 1
    last_sel = row_n - 1
    add_tenho_validation(ws, "H", 2, last_sel)
    add_tenho_conditional(ws, "H", 2, last_sel)
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "E2"
    # Não usar worksheet.auto_filter junto com Table — Excel/openpyxl corrompem o xlsx.
    attach_table(ws, f"A1:I{last_sel}", "tblSelecoes", stripe=True)

    # --- Museu FIFA ---
    wm = wb.create_sheet("Museu FIFA")
    hm = ["#", "Código", "Descrição no álbum", "Tipo", "Tenho", "Duplicatas"]
    for c, h in enumerate(hm, 1):
        wm.cell(row=1, column=c, value=h)
    apply_header(wm, 1, len(hm))
    for i, (code, desc) in enumerate(museum, start=1):
        wm.cell(row=i + 1, column=1, value=i)
        wm.cell(row=i + 1, column=2, value=code)
        wm.cell(row=i + 1, column=3, value=desc)
        wm.cell(row=i + 1, column=4, value=foil_type(desc))
        wm.cell(row=i + 1, column=6, value=0)
        wm.row_dimensions[i + 1].height = 22
    last_mu = 1 + len(museum)
    add_tenho_validation(wm, "E", 2, last_mu)
    add_tenho_conditional(wm, "E", 2, last_mu)
    wm.column_dimensions["E"].width = 16
    wm.column_dimensions["C"].width = 40
    wm.freeze_panes = "E2"
    attach_table(wm, f"A1:F{last_mu}", "tblMuseu", stripe=True)

    # --- Coca-Cola ---
    wc = wb.create_sheet("Coca-Cola")
    ch = ["# Promo", "Jogador", "Seleção", "Origem", "Tenho", "Duplicatas"]
    for c, h in enumerate(ch, 1):
        wc.cell(row=1, column=c, value=h)
    apply_header(wc, 1, len(ch))
    for i, (num, player, sel) in enumerate(COCA_COLA, start=2):
        wc.cell(row=i, column=1, value=num)
        wc.cell(row=i, column=2, value=player)
        wc.cell(row=i, column=3, value=sel)
        wc.cell(row=i, column=4, value="Rótulo Coca-Cola EUA")
        wc.cell(row=i, column=6, value=0)
        wc.row_dimensions[i].height = 22
    last_cc = 1 + len(COCA_COLA)
    add_tenho_validation(wc, "E", 2, last_cc)
    add_tenho_conditional(wc, "E", 2, last_cc)
    wc.column_dimensions["E"].width = 16
    wc.freeze_panes = "E2"
    attach_table(wc, f"A1:F{last_cc}", "tblCocaCola", stripe=True)

    # --- Extras ---
    we = wb.create_sheet("Extras")
    eh = ["#", "Jogador", "Seleção", "Notas", "Tenho", "Duplicatas"]
    for c, h in enumerate(eh, 1):
        we.cell(row=1, column=c, value=h)
    apply_header(we, 1, len(eh))
    for i, (player, sel) in enumerate(EXTRAS, start=2):
        we.cell(row=i, column=1, value=i - 1)
        we.cell(row=i, column=2, value=player)
        we.cell(row=i, column=3, value=sel)
        we.cell(row=i, column=4, value="Extra internacional (sem nº no álbum)")
        we.cell(row=i, column=6, value=0)
        we.row_dimensions[i].height = 22
    last_ex = 1 + len(EXTRAS)
    add_tenho_validation(we, "E", 2, last_ex)
    add_tenho_conditional(we, "E", 2, last_ex)
    we.column_dimensions["E"].width = 16
    we.freeze_panes = "E2"
    attach_table(we, f"A1:F{last_ex}", "tblExtras", stripe=True)

    # --- Siglas & Grupos (referência rápida) ---
    wx = wb.create_sheet("Siglas")
    xh = ["Sigla FIFA", "Nome da seleção", "Grupo", "Figurinhas", "Obs."]
    for c, h in enumerate(xh, 1):
        wx.cell(row=1, column=c, value=h)
    apply_header(wx, 1, len(xh))
    for i, tm in enumerate(teams_meta, start=2):
        wx.cell(row=i, column=1, value=tm["sigla"])
        wx.cell(row=i, column=2, value=tm["pais"])
        wx.cell(row=i, column=3, value=tm["grupo"])
        wx.cell(row=i, column=4, value=20)
        if tm["sigla_panini"] != tm["sigla"]:
            wx.cell(row=i, column=5, value=f"Panini: «{tm['sigla_panini']}»")
    wx.freeze_panes = "A2"
    wx.auto_filter.ref = wx.dimensions
    autosize_columns(wx)

    wg = wb.create_sheet("Grupos")
    gh = ["Grupo", "Seleção 1", "Seleção 2", "Seleção 3", "Seleção 4"]
    for c, h in enumerate(gh, 1):
        wg.cell(row=1, column=c, value=h)
    apply_header(wg, 1, len(gh))
    for gi in range(12):
        slice_t = teams_meta[gi * 4 : (gi + 1) * 4]
        row = gi + 2
        wg.cell(row=row, column=1, value=f"Grupo {chr(ord('A') + gi)}")
        for j, tm in enumerate(slice_t):
            wg.cell(row=row, column=2 + j, value=f"{tm['pais']} ({tm['sigla']})")
    wg.freeze_panes = "A2"
    autosize_columns(wg)

    # --- Lista única (980) — impressão / cópia rápida ---
    wa = wb.create_sheet("Todas (980)")
    ah = ["Secção", "Grupo", "Sigla", "Seleção", "Código", "Descrição", "Tipo", "Tenho", "Dup."]
    for c, h in enumerate(ah, 1):
        wa.cell(row=1, column=c, value=h)
    apply_header(wa, 1, len(ah))
    rr = 2
    for code, desc in intro:
        wa.cell(row=rr, column=1, value="Página inicial")
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    for tm in teams_meta:
        for code, desc in tm["block"]:
            wa.cell(row=rr, column=1, value="Seleção")
            wa.cell(row=rr, column=2, value=tm["grupo"])
            wa.cell(row=rr, column=3, value=tm["sigla"])
            wa.cell(row=rr, column=4, value=tm["pais"])
            wa.cell(row=rr, column=5, value=code)
            wa.cell(row=rr, column=6, value=desc)
            wa.cell(row=rr, column=7, value=foil_type(desc))
            wa.cell(row=rr, column=9, value=0)
            rr += 1
    for code, desc in museum:
        wa.cell(row=rr, column=1, value="Museu FIFA")
        wa.cell(row=rr, column=5, value=code)
        wa.cell(row=rr, column=6, value=desc)
        wa.cell(row=rr, column=7, value=foil_type(desc))
        wa.cell(row=rr, column=9, value=0)
        rr += 1
    add_tenho_validation(wa, "H", 2, rr - 1)
    add_tenho_conditional(wa, "H", 2, rr - 1)
    wa.freeze_panes = "E2"
    wa.column_dimensions["H"].width = 16
    attach_table(wa, f"A1:I{rr - 1}", "tblTodas980", stripe=True)
    autosize_columns(wa)

    meta["last_pi"] = last_pi
    meta["last_sel"] = last_sel
    meta["last_mu"] = last_mu
    meta["last_cc"] = last_cc
    meta["last_ex"] = last_ex
    return meta


def build_estatisticas(wb: Workbook, teams_meta) -> dict:
    """Motor numérico + tabela por grupo (alimenta gráficos)."""
    ws = wb.create_sheet("Estatísticas")
    ws.sheet_properties.tabColor = "06D6A0"

    ws["A1"] = "Panini WC 2026 — Motor de estatísticas"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws.merge_cells("A1:H1")

    # --- Bloco: totais por secção ---
    ws["A3"] = "Secção"
    ws["B3"] = "Meta"
    ws["C3"] = "Com «Sim»"
    ws["D3"] = "%"
    ws["E3"] = "Faltam"
    for c in range(1, 6):
        ws.cell(row=3, column=c).font = Font(bold=True)
        ws.cell(row=3, column=c).fill = PatternFill("solid", fgColor="E8EEF7")

    rows_sec = [
        ("Página inicial", 9, "COUNTIF(tblPaginaInicial[Tenho],\"Sim\")"),
        ("Seleções (48×20)", 960, "COUNTIF(tblSelecoes[Tenho],\"Sim\")"),
        ("Museu FIFA", 11, "COUNTIF(tblMuseu[Tenho],\"Sim\")"),
    ]
    r = 4
    for label, meta_val, formula_sim in rows_sec:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=meta_val)
        ws.cell(row=r, column=3, value=f"={formula_sim}")
        ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
        ws.cell(row=r, column=5, value=f"=B{r}-C{r}")
        ws.cell(row=r, column=4).number_format = "0.0%"
        r += 1

    ws.cell(row=r, column=1, value="ÁLBUM BASE (total)")
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=2, value=980)
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(
        row=r,
        column=3,
        value=(
            "=COUNTIF(tblPaginaInicial[Tenho],\"Sim\")"
            "+COUNTIF(tblSelecoes[Tenho],\"Sim\")"
            "+COUNTIF(tblMuseu[Tenho],\"Sim\")"
        ),
    )
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=5, value=f"=B{r}-C{r}")
    ws.cell(row=r, column=4).number_format = "0.00%"
    row_album_total = r

    r += 2
    ws.cell(row=r, column=1, value="Promo & extras")
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value="Item")
    ws.cell(row=r, column=2, value="Meta")
    ws.cell(row=r, column=3, value="Com «Sim»")
    ws.cell(row=r, column=4, value="%")
    for c in range(1, 5):
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Coca-Cola")
    ws.cell(row=r, column=2, value=len(COCA_COLA))
    ws.cell(row=r, column=3, value='=COUNTIF(tblCocaCola[Tenho],"Sim")')
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=4).number_format = "0%"
    r += 1
    ws.cell(row=r, column=1, value="Extras internacionais")
    ws.cell(row=r, column=2, value=len(EXTRAS))
    ws.cell(row=r, column=3, value='=COUNTIF(tblExtras[Tenho],"Sim")')
    ws.cell(row=r, column=4, value=f"=IF(B{r}>0,C{r}/B{r},0)")
    ws.cell(row=r, column=4).number_format = "0%"
    row_ext_end = r

    # --- FOIL vs Base (álbum base apenas em tblSelecoes + intro + museu) ---
    r = row_ext_end + 3
    ws.cell(row=r, column=1, value="Análise FOIL (onde marcou «Sim»)")
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value="FOIL com Sim (todas as folhas base)")
    ws.cell(
        row=r,
        column=2,
        value=(
            "=SUMPRODUCT((tblPaginaInicial[Tipo]=\"FOIL\")*(tblPaginaInicial[Tenho]=\"Sim\"))"
            "+SUMPRODUCT((tblSelecoes[Tipo]=\"FOIL\")*(tblSelecoes[Tenho]=\"Sim\"))"
            "+SUMPRODUCT((tblMuseu[Tipo]=\"FOIL\")*(tblMuseu[Tenho]=\"Sim\"))"
        ),
    )
    r += 1
    ws.cell(row=r, column=1, value="Base (não foil) com Sim")
    ws.cell(
        row=r,
        column=2,
        value=(
            "=SUMPRODUCT((tblPaginaInicial[Tipo]=\"Base\")*(tblPaginaInicial[Tenho]=\"Sim\"))"
            "+SUMPRODUCT((tblSelecoes[Tipo]=\"Base\")*(tblSelecoes[Tenho]=\"Sim\"))"
            "+SUMPRODUCT((tblMuseu[Tipo]=\"Base\")*(tblMuseu[Tenho]=\"Sim\"))"
        ),
    )
    row_foil_end = r

    # --- Por grupo (12 linhas) — referência para gráfico de barras ---
    r = row_foil_end + 3
    ws.cell(row=r, column=1, value="Progresso por grupo (seleções)")
    ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    r += 1
    grp_header_row = r
    ws.cell(row=r, column=1, value="Grupo")
    ws.cell(row=r, column=2, value="Meta")
    ws.cell(row=r, column=3, value="Com Sim")
    ws.cell(row=r, column=4, value="% completo")
    ws.cell(row=r, column=5, value="Faltam")
    ws.cell(row=r, column=6, value="Barra visual")
    for c in range(1, 7):
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="D6E4FF")
    r += 1
    grp_first_row = r

    for gi in range(12):
        gname = f"Grupo {chr(ord('A') + gi)}"
        row = r + gi
        ws.cell(row=row, column=1, value=gname)
        ws.cell(row=row, column=2, value=80)
        ws.cell(
            row=row,
            column=3,
            value=f'=COUNTIFS(tblSelecoes[Tenho],"Sim",tblSelecoes[Grupo],"{gname}")',
        )
        ws.cell(row=row, column=4, value=f"=IF(B{row}>0,C{row}/B{row},0)")
        ws.cell(row=row, column=5, value=f"=B{row}-C{row}")
        ws.cell(row=row, column=4).number_format = "0%"
        ws.cell(
            row=row,
            column=6,
            value=f'=REPT("█",ROUND(D{row}*12,0))&REPT("·",12-ROUND(D{row}*12,0))',
        )
    grp_last_row = r + 11
    attach_table(ws, f"A{grp_header_row}:F{grp_last_row}", "tblPorGrupo", stripe=True)

    # Duplicatas totais
    r = grp_last_row + 3
    ws.cell(row=r, column=1, value="Soma de duplicatas (pacotes extras)")
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(
        row=r,
        column=2,
        value=(
            "=SUM(tblPaginaInicial[Duplicatas])+SUM(tblSelecoes[Duplicatas])"
            "+SUM(tblMuseu[Duplicatas])+SUM(tblCocaCola[Duplicatas])+SUM(tblExtras[Duplicatas])"
        ),
    )

    autosize_columns(ws)

    return {
        "row_album_total": row_album_total,
        "grp_first_row": grp_first_row,
        "grp_last_row": grp_last_row,
        "grp_header_row": grp_header_row,
    }


def build_dashboard(wb: Workbook, stats: dict) -> None:
    wd = wb["Dashboard"]
    rt = stats["row_album_total"]
    gf = stats["grp_first_row"]
    gl = stats["grp_last_row"]

    for row in range(1, 45):
        for col in range(1, 22):
            wd.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BG_DASH)
    wd.sheet_view.showGridLines = False

    wd.merge_cells("A1:T2")
    c = wd["A1"]
    c.value = "Copa do Mundo 2026 · Panini — Dashboard"
    c.font = DASH_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center")

    wd.merge_cells("A3:T3")
    wd["A3"] = (
        "Preencha «Tenho» nas folhas «Página Inicial», «Seleções», «Museu FIFA», "
        "«Coca-Cola» e «Extras». KPIs e gráficos atualizam automaticamente."
    )
    wd["A3"].font = DASH_SUB
    wd["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Título e valor em merges separados (linha a linha) para evitar MergedCell só-leitura.
    cards = [
        ("B5:D5", "B6:D6", "Álbum base (coladas)", f"=Estatísticas!C{rt}&\" / \"&Estatísticas!B{rt}"),
        ("E5:G5", "E6:G6", "Faltam (base)", f"=Estatísticas!E{rt}"),
        ("H5:J5", "H6:J6", "% Completo", f"=Estatísticas!D{rt}"),
        ("K5:M5", "K6:M6", "Coca-Cola", "=COUNTIF(tblCocaCola[Tenho],\"Sim\")&\" / \"&COUNTA(tblCocaCola[Tenho])"),
        ("N5:P5", "N6:P6", "Extras", "=COUNTIF(tblExtras[Tenho],\"Sim\")&\" / \"&COUNTA(tblExtras[Tenho])"),
        ("B8:D8", "B9:D9", "Duplicatas (todas)", "=SUM(tblPaginaInicial[Duplicatas])+SUM(tblSelecoes[Duplicatas])+SUM(tblMuseu[Duplicatas])+SUM(tblCocaCola[Duplicatas])+SUM(tblExtras[Duplicatas])"),
        ("E8:G8", "E9:G9", "«Falta trocar»", "=COUNTIF(tblPaginaInicial[Tenho],\"Falta trocar\")+COUNTIF(tblSelecoes[Tenho],\"Falta trocar\")+COUNTIF(tblMuseu[Tenho],\"Falta trocar\")"),
        ("H8:J8", "H9:J9", "FOIL coladas (Sim)", "=SUMPRODUCT((tblPaginaInicial[Tipo]=\"FOIL\")*(tblPaginaInicial[Tenho]=\"Sim\"))+SUMPRODUCT((tblSelecoes[Tipo]=\"FOIL\")*(tblSelecoes[Tenho]=\"Sim\"))+SUMPRODUCT((tblMuseu[Tipo]=\"FOIL\")*(tblMuseu[Tenho]=\"Sim\"))"),
        ("K8:M8", "K9:M9", "Data", "=TODAY()"),
    ]

    for rng_title, rng_val, title, formula in cards:
        wd.merge_cells(rng_title)
        wd.merge_cells(rng_val)
        tl = rng_title.split(":")[0]
        vl = rng_val.split(":")[0]
        wd[tl].value = title
        wd[tl].font = DASH_LBL
        wd[tl].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cval = wd[vl]
        cval.value = formula
        cval.alignment = Alignment(horizontal="center", vertical="center")
        if "% Completo" in title:
            cval.font = DASH_KPI
            cval.number_format = "0.0%"
        elif title == "Data":
            cval.font = Font(size=14, color="FFFFFF", bold=True)
            cval.number_format = "dddd dd/mm/yyyy"
        else:
            cval.font = Font(size=18, color="FFFFFF", bold=True)

    # Área auxiliar para gráficos (fora dos merges A1:T3)
    wd["R40"] = "Coladas"
    wd["S40"] = f"=Estatísticas!C{rt}"
    wd["R41"] = "Faltam"
    wd["S41"] = f"=Estatísticas!E{rt}"

    dough = DoughnutChart()
    dough.title = "Álbum base · Coladas vs faltam"
    dough.style = 26
    data_d = Reference(wd, min_col=19, min_row=40, max_row=41)
    cats_d = Reference(wd, min_col=18, min_row=40, max_row=41)
    dough.add_data(data_d, titles_from_data=False)
    dough.set_categories(cats_d)
    dough.dataLabels = DataLabelList()
    dough.dataLabels.showPercent = True
    wd.add_chart(dough, "B12")
    wd.row_dimensions[40].hidden = True
    wd.row_dimensions[41].hidden = True

    bars = BarChart()
    bars.type = "bar"
    bars.style = 10
    bars.title = "Progresso por grupo (seleções)"
    bars.y_axis.title = None
    bars.x_axis.title = "% completo"
    data_b = Reference(wb["Estatísticas"], min_col=4, min_row=gf, max_row=gl)
    cats_b = Reference(wb["Estatísticas"], min_col=1, min_row=gf, max_row=gl)
    bars.add_data(data_b, titles_from_data=False)
    bars.set_categories(cats_b)
    bars.shape = 4
    wd.add_chart(bars, "K12")

    wd.merge_cells("B34:T38")
    wd["B34"] = (
        "Dica: ordene ou filtre «Seleções» pela coluna «Tenho»; "
        "na «Estatísticas» encontra detalhe por secção, FOIL/Base e por grupo. "
        "Abas com cor na etiqueta: motor de cálculo."
    )
    wd["B34"].font = Font(size=10, color="B8C5D9")
    wd["B34"].alignment = Alignment(wrap_text=True, vertical="top")

    wd.column_dimensions["B"].width = 14
    wd.sheet_properties.tabColor = "3A86FF"


def main() -> None:
    intro, museum, teams_meta = load_album(CHECKLIST_DEFAULT)
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Dashboard", 0)

    build_data_sheets(wb, intro, museum, teams_meta)
    stats = build_estatisticas(wb, teams_meta)
    build_dashboard(wb, stats)

    wb.save(OUT_V2)
    print(f"Escrito: {OUT_V2}")


if __name__ == "__main__":
    main()
